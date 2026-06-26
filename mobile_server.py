from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import threading
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse, JSONResponse

BASE_DIR = Path(__file__).resolve().parent
STATUS_PATH = BASE_DIR / "status.json"
CONFIG_PATH = BASE_DIR / "config.json"
REPORT_DIR = BASE_DIR / "Reports"
LOG_PATH = BASE_DIR / "logs" / "mobile_task.log"
HISTORY_PATH = BASE_DIR / "logs" / "run_history.json"
REAL_RUN_LIMIT_PATH = BASE_DIR / "data" / "real_run_limit.json"
RUNTIME_MODE_PATH = BASE_DIR / "data" / "runtime_mode.json"
PRODUCTION_HISTORY_PATH = BASE_DIR / "data" / "production_run_history.json"
ALLOWED_STATES = {"idle", "running", "stopping", "completed", "failed", "interrupted", "stopped", "PRECHECK_FAILED"}
APP_VERSION = "2.0.0-stable"
SERVICE_STARTED_AT = datetime.now().isoformat(timespec="seconds")
SENSITIVE_RE = re.compile(r"cookie|token|session|password|secret", re.I)
COMPLETED_ACCOUNT_STATES = {"SUCCESS", "PARTIAL_SUCCESS", "NO_ORDERS"}


class Utf8JSONResponse(JSONResponse):
    media_type = "application/json; charset=utf-8"


app = FastAPI(title="CrewSupply Mobile Control", default_response_class=Utf8JSONResponse)
_lock = threading.Lock()
_process: subprocess.Popen | None = None


def read_config() -> Dict[str, Any]:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8-sig"))


def mask_path(value: str) -> str:
    if not value:
        return ""
    try:
        path = Path(value)
        if path.is_absolute():
            return str(path.resolve().relative_to(BASE_DIR)).replace("\\", "/")
    except Exception:
        pass
    return str(value)


def sanitize_value(value: Any) -> Any:
    if isinstance(value, dict):
        return {k: ("***" if SENSITIVE_RE.search(k) else sanitize_value(v)) for k, v in value.items() if not SENSITIVE_RE.search(k)}
    if isinstance(value, list):
        return [sanitize_value(item) for item in value]
    if isinstance(value, str):
        if SENSITIVE_RE.search(value):
            return "***"
        return mask_path(value)
    return value


def read_json(path: Path, fallback: Any) -> Any:
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:
            return fallback
    return fallback


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def pid_is_alive(pid: Any) -> bool:
    try:
        pid = int(pid)
    except (TypeError, ValueError):
        return False
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
        return True
    except PermissionError:
        return True
    except OSError:
        return False


def reconcile_stale_status() -> Dict[str, Any]:
    value = read_json(STATUS_PATH, {"state": "idle", "message": "尚未运行任务"})
    if value.get("state") != "running":
        return value
    pid = value.get("pid")
    updated = _parse_dt(str(value.get("updated_at", "")))
    stale_without_pid = pid is None and updated is not None and (datetime.now() - updated).total_seconds() > 600
    dead_pid = pid is not None and not pid_is_alive(pid)
    if dead_pid or stale_without_pid:
        value.update({
            "state": "interrupted",
            "message": "上次任务已中断",
            "current_step": "任务进程已结束",
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        })
        write_json(STATUS_PATH, value)
    return value


def find_latest_report() -> Path | None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    reports = [p for p in REPORT_DIR.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    return max(reports, key=lambda p: p.stat().st_mtime) if reports else None


def report_summary(report_path: Path) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}
    if not report_path.exists():
        return {}
    try:
        wb = load_workbook(report_path, read_only=True, data_only=True)
    except Exception:
        return {}

    def sheet_rows(index: int) -> List[tuple]:
        if len(wb.worksheets) <= index:
            return []
        rows = []
        for row in wb.worksheets[index].iter_rows(min_row=2, values_only=True):
            if any(value is not None for value in row):
                rows.append(row)
        return rows

    shipping = sheet_rows(0)
    errors = sheet_rows(1)
    summary_rows = sheet_rows(2)
    carriers = Counter(str(row[7] or "").upper() for row in shipping if len(row) > 7)
    account_statuses = {}
    for row in summary_rows:
        if len(row) >= 2 and row[0]:
            account_statuses[str(row[0])] = str(row[1] or "")
    duplicate_skipped = sum(int(row[8] or 0) for row in summary_rows if len(row) > 8 and str(row[8] or "0").isdigit())
    result = {
        "successful_pdfs": len(shipping),
        "success_pdf_count": len(shipping),
        "exceptions": len(errors),
        "exception_count": len(errors),
        "duplicate_skipped_count": duplicate_skipped,
        "ups_count": carriers.get("UPS", 0),
        "usps_count": carriers.get("USPS", 0),
        "fedex_count": carriers.get("FEDEX", 0),
        "unknown_count": carriers.get("UNKNOWN", 0),
        "account_statuses": account_statuses,
    }
    if "account_YOUT" in account_statuses:
        result["account_YOUT"] = account_statuses["account_YOUT"]
    try:
        wb.close()
    except Exception:
        pass
    return result


def enrich_status_from_latest_result(value: Dict[str, Any]) -> Dict[str, Any]:
    report_value = value.get("report") or value.get("excel") or ""
    report = Path(report_value) if report_value else None
    if report is not None and not report.is_absolute():
        report = BASE_DIR / report
    may_use_latest_report = (
        value.get("state") == "completed"
        and not report_value
        and "recovery" in str(value.get("run_id", "")).lower()
    )
    if (report is None or not report.exists()) and may_use_latest_report:
        report = find_latest_report()
    if report is not None and report.exists():
        summary = report_summary(report)
        if summary:
            value.update(summary)
            value["report"] = str(report)
    value.setdefault("current_step", value.get("message", ""))
    if "processed_success_count" in value and not value.get("success_pdf_count"):
        value["success_pdf_count"] = value.get("processed_success_count", 0)
        value["successful_pdfs"] = value.get("processed_success_count", 0)
    if "hold_count" in value and not value.get("exception_count"):
        value["exception_count"] = value.get("hold_count", 0)
        value["exceptions"] = value.get("hold_count", 0)
    return value


def current_status() -> Dict[str, Any]:
    value = reconcile_stale_status()
    original = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    value = enrich_status_from_latest_result(value)
    if json.dumps(value, ensure_ascii=False, sort_keys=True, default=str) != original:
        write_json(STATUS_PATH, value)
    state = value.get("state", "idle")
    if state not in ALLOWED_STATES:
        state = "failed" if state == "stopped" else "idle"
    sanitized = sanitize_value(value)
    sanitized["state"] = state
    sanitized.setdefault("current_account", sanitized.get("account", ""))
    sanitized.setdefault("current_step", sanitized.get("message", ""))
    sanitized.setdefault("completed_accounts", sanitized.get("current", 0))
    sanitized.setdefault("success_pdf_count", sanitized.get("successful_pdfs", sanitized.get("processed_success_count", sanitized.get("success", 0))))
    sanitized.setdefault("successful_pdfs", sanitized.get("success_pdf_count", 0))
    sanitized.setdefault("no_order_count", sanitized.get("no_orders", 0))
    sanitized.setdefault("exception_count", sanitized.get("exceptions", sanitized.get("hold_count", 0)))
    sanitized.setdefault("exceptions", sanitized.get("exception_count", 0))
    sanitized.setdefault("duplicate_skipped_count", sanitized.get("duplicates", 0))
    sanitized.setdefault("started_at", sanitized.get("start_time", ""))
    sanitized.setdefault("updated_at", datetime.now().isoformat(timespec="seconds"))
    if _process is not None:
        sanitized["process_running"] = _process.poll() is None
    return sanitized


def latest_file(pattern: str) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(
        (p for p in REPORT_DIR.glob(pattern) if p.is_file() and not p.name.startswith("~$")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not files:
        raise HTTPException(status_code=404, detail="No output file yet")
    return safe_report_path(files[0].name, pattern)


def safe_report_path(filename: str, pattern: str | None = None) -> Path:
    if ".." in Path(filename).parts or Path(filename).is_absolute():
        raise HTTPException(status_code=400, detail="Invalid download path")
    path = (REPORT_DIR / filename).resolve()
    root = REPORT_DIR.resolve()
    if root != path and root not in path.parents:
        raise HTTPException(status_code=403, detail="Download path is outside Reports")
    if pattern == "*.xlsx" and path.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")
    if pattern == "*.zip" and path.suffix.lower() != ".zip":
        raise HTTPException(status_code=400, detail="Only .zip files are allowed")
    if not path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    return path


def append_history_from_status(status: Dict[str, Any]) -> None:
    if status.get("state") not in {"completed", "failed"}:
        return
    history: List[Dict[str, Any]] = read_json(HISTORY_PATH, [])
    run_id = status.get("run_id") or status.get("updated_at") or datetime.now().isoformat(timespec="seconds")
    if any(item.get("run_id") == run_id for item in history):
        return
    history.insert(0, {
        "run_id": run_id,
        "time": status.get("updated_at", ""),
        "state": status.get("state", ""),
        "success": status.get("success_pdf_count", status.get("success", 0)),
        "no_orders": status.get("no_order_count", status.get("no_orders", 0)),
        "exceptions": status.get("exception_count", status.get("exceptions", 0)),
        "excel": Path(status.get("report", "")).name if status.get("report") else "",
        "zip": Path(status.get("archive", "")).name if status.get("archive") else "",
    })
    write_json(HISTORY_PATH, history[:10])


def real_runs_available() -> bool:
    payload = read_json(REAL_RUN_LIMIT_PATH, {"allowed_real_runs": 2, "completed_real_runs": 0})
    return int(payload.get("completed_real_runs", 0)) < int(payload.get("allowed_real_runs", 2))


def task_is_running() -> bool:
    if _process is not None and _process.poll() is None:
        return True
    status = reconcile_stale_status()
    if status.get("state") != "running":
        return False
    pid = status.get("pid")
    if pid is not None:
        return pid_is_alive(pid)
    # Backward-compatible for tests or a just-started legacy status without a pid.
    updated = _parse_dt(str(status.get("updated_at", "")))
    return updated is None or (datetime.now() - updated).total_seconds() <= 600


def account_completion_plan() -> Dict[str, Any]:
    report = find_latest_report()
    statuses: Dict[str, str] = {}
    if report is not None:
        summary = report_summary(report)
        statuses.update(summary.get("account_statuses", {}))
    accounts = [account["name"] for account in read_config().get("accounts", [])]
    normalized = {name: statuses.get(name, "MISSING") for name in accounts}
    to_run = [name for name in accounts if normalized.get(name) not in COMPLETED_ACCOUNT_STATES]
    failed = [name for name in to_run if normalized.get(name) == "FAILED"]
    return {
        "latest_report": str(report) if report else "",
        "account_statuses": normalized,
        "to_run": to_run,
        "pending_accounts": to_run,
        "failed_accounts": failed,
        "skipped": [name for name in accounts if name not in to_run],
        "running": task_is_running(),
        "blocked_reason": "请先停止当前任务" if task_is_running() else "",
    }


def default_runtime_mode() -> Dict[str, Any]:
    return {
        "mode": "production",
        "production_enabled": True,
        "allow_mobile_run": True,
        "max_runs_per_day": 1,
        "require_confirmation": True,
        "show_account_debug": False,
    }


def runtime_mode() -> Dict[str, Any]:
    payload = read_json(RUNTIME_MODE_PATH, default_runtime_mode())
    merged = default_runtime_mode()
    if isinstance(payload, dict):
        merged.update(payload)
    return merged


def production_mode_enabled() -> bool:
    mode = runtime_mode()
    return bool(mode.get("production_enabled")) and mode.get("mode") == "production" and bool(mode.get("allow_mobile_run"))


def today_key() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def read_production_history() -> Dict[str, Any]:
    payload = read_json(PRODUCTION_HISTORY_PATH, {"runs": [], "manual_resets": []})
    if not isinstance(payload, dict):
        payload = {"runs": [], "manual_resets": []}
    payload.setdefault("runs", [])
    payload.setdefault("manual_resets", [])
    return payload


def write_production_history(payload: Dict[str, Any]) -> None:
    payload.setdefault("runs", [])
    payload.setdefault("manual_resets", [])
    write_json(PRODUCTION_HISTORY_PATH, payload)


def _parse_dt(value: str) -> datetime | None:
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None


def latest_manual_reset_at(history: Dict[str, Any], run_date: str) -> datetime | None:
    values = []
    for reset in history.get("manual_resets", []):
        if reset.get("date") == run_date:
            parsed = _parse_dt(str(reset.get("time", "")))
            if parsed is not None:
                values.append(parsed)
    return max(values) if values else None


def production_runs_today(history: Dict[str, Any] | None = None) -> List[Dict[str, Any]]:
    history = history or read_production_history()
    run_date = today_key()
    reset_at = latest_manual_reset_at(history, run_date)
    runs = []
    for item in history.get("runs", []):
        if item.get("date") != run_date or item.get("scope") != "all_accounts":
            continue
        started = _parse_dt(str(item.get("started_at", "")))
        if reset_at is not None and started is not None and started <= reset_at:
            continue
        runs.append(item)
    return runs


def can_start_production_run() -> bool:
    if not production_mode_enabled():
        return False
    max_runs = int(runtime_mode().get("max_runs_per_day") or 1)
    return len(production_runs_today()) < max_runs


def reserve_production_run(command: List[str]) -> Dict[str, Any]:
    if not can_start_production_run():
        raise HTTPException(status_code=403, detail="今日正式运行已执行，不能重复启动")
    history = read_production_history()
    now = datetime.now().isoformat(timespec="seconds")
    entry = {
        "date": today_key(),
        "run_id": f"pending_{datetime.now():%Y%m%d_%H%M%S}",
        "scope": "all_accounts",
        "status": "RUNNING",
        "started_at": now,
        "ended_at": "",
        "account_statuses": {},
        "excel_path": "",
        "pdf_zip_path": "",
        "command": " ".join(command),
    }
    history.setdefault("runs", []).append(entry)
    write_production_history(history)
    return entry


def update_latest_production_run(status: Dict[str, Any]) -> None:
    if status.get("state") not in {"completed", "failed", "PRECHECK_FAILED"}:
        return
    history = read_production_history()
    today_runs = [r for r in history.get("runs", []) if r.get("date") == today_key() and r.get("scope") == "all_accounts"]
    if not today_runs:
        return
    entry = today_runs[-1]
    if entry.get("ended_at"):
        return
    entry["status"] = str(status.get("state", "")).upper()
    entry["run_id"] = str(status.get("run_id") or entry.get("run_id") or "")
    entry["ended_at"] = datetime.now().isoformat(timespec="seconds")
    entry["account_statuses"] = sanitize_value(status.get("account_statuses", status.get("accounts", {})))
    entry["excel_path"] = mask_path(str(status.get("report", "")))
    entry["pdf_zip_path"] = mask_path(str(status.get("archive", "")))
    write_production_history(history)


def reset_production_today(reason: str = "manual") -> Dict[str, Any]:
    history = read_production_history()
    entry = {"date": today_key(), "time": datetime.now().isoformat(timespec="seconds"), "reason": reason or "manual"}
    history.setdefault("manual_resets", []).append(entry)
    write_production_history(history)
    return entry


def _watch_process(proc: subprocess.Popen) -> None:
    global _process
    return_code = proc.wait()
    time.sleep(0.2)
    with _lock:
        if _process is proc:
            _process = None
    status = read_json(STATUS_PATH, {})
    if status.get("state") in {"running", "stopping"}:
        status.update({
            "state": "interrupted" if return_code == 0 else "failed",
            "message": "任务进程已结束但未写入最终状态" if return_code == 0 else f"任务进程异常退出，代码 {return_code}",
            "current_step": "进程已结束",
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "exit_code": return_code,
        })
        write_json(STATUS_PATH, status)


def launch(args: List[str], production: bool = False) -> Dict[str, Any]:
    global _process
    command = [sys.executable, str(BASE_DIR / "auto_all_integrated.py"), *args]
    is_real_run = "--demo" not in args and "--process-only" not in args and "--resume-zip" not in args and "--production-run" not in args
    with _lock:
        if _process is not None and _process.poll() is None:
            raise HTTPException(status_code=409, detail="A task is already running")
        if production:
            reserve_production_run(command)
        elif is_real_run and not real_runs_available():
            raise HTTPException(status_code=403, detail="Real run limit reached; third real run is forbidden")
        LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        status = {
            "state": "running",
            "message": "任务已启动",
            "current_step": "启动中",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "production": production,
            "command_scope": " ".join(args),
        }
        write_json(STATUS_PATH, status)
        log_file = LOG_PATH.open("a", encoding="utf-8")
        try:
            _process = subprocess.Popen(command, cwd=str(BASE_DIR), stdout=log_file, stderr=subprocess.STDOUT)
            status["pid"] = _process.pid
            write_json(STATUS_PATH, status)
            if _process.__class__.__module__ == "subprocess":
                threading.Thread(target=_watch_process, args=(_process,), daemon=True).start()
        finally:
            log_file.close()
    return {"ok": True, "pid": _process.pid, "production": production}


def safe_log_tail(lines: int = 80) -> str:
    if not LOG_PATH.exists():
        return ""
    content = LOG_PATH.read_text(encoding="utf-8", errors="ignore").splitlines()[-lines:]
    sanitized = []
    for line in content:
        sanitized.append("***" if SENSITIVE_RE.search(line) else mask_path(line))
    return "\n".join(sanitized)


@app.get("/")
def home():
    mode = runtime_mode()
    production = production_mode_enabled()
    can_run = can_start_production_run() if production else real_runs_available()
    disabled = "" if can_run else " disabled"
    start_label = "开始今日发货" if production else "运行全部账号"
    run_url = "/run/all"
    accounts = "" if not mode.get("show_account_debug", False) else "".join(
        f'<button class="secondary" onclick="post(\'/run/account/{a["name"]}\')">单独运行 {a["name"]}</button>'
        for a in read_config()["accounts"]
    )
    confirm_text = "确认运行今天的全部账号吗？每个账号只执行一次AWB生成。"
    pending_plan = account_completion_plan()
    pending_disabled = " disabled" if pending_plan.get("running") or not pending_plan.get("to_run") else ""
    pending_title = pending_plan.get("blocked_reason") or ("今天没有需要补跑的账号" if not pending_plan.get("to_run") else "")
    html = f"""
<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CrewSupply 手机控制中心</title>
<style>
body{{font-family:Arial,sans-serif;margin:0;background:#f4f6f8;color:#1f2933}}main{{max-width:720px;margin:0 auto;padding:14px}}
.card{{background:white;padding:14px;border-radius:8px;margin-bottom:12px;box-shadow:0 1px 6px #0001}}button,a{{display:block;width:100%;box-sizing:border-box;margin:8px 0;padding:13px;border:0;border-radius:8px;background:#1f4e78;color:white;text-align:center;text-decoration:none;font-size:16px}}
button:disabled{{background:#98a2b3;color:#eef2f5}}input{{display:block;width:100%;box-sizing:border-box;margin:8px 0;padding:12px;border:1px solid #cbd5e1;border-radius:8px;font-size:15px}}
.secondary{{background:#526d82}}.grid{{display:grid;grid-template-columns:1fr 1fr;gap:8px}}.label{{font-size:12px;color:#667085}}.value{{font-weight:700;word-break:break-word}}pre{{white-space:pre-wrap;word-break:break-word;background:#eef2f5;padding:10px;border-radius:8px;max-height:260px;overflow:auto}}
</style>
</head>
<body><main>
<div class="card"><h2>CrewSupply 手机控制中心</h2><button id="startBtn"{disabled} onclick="confirmPost('{run_url}')">{start_label}</button><button class="secondary" onclick="loadAll()">查看状态</button></div>
<div class="card"><h3>文件下载</h3><a href="/download/latest-excel">下载最新Excel</a><a href="/download/latest-zip">下载最新PDF ZIP</a></div>
<div class="card"><details><summary>高级工具</summary>{accounts}<input id="resumeZip" placeholder="本地ZIP完整路径"><button class="secondary" onclick="resumeZip()">从本地ZIP继续处理</button><button id="pendingBtn" class="secondary"{pending_disabled} title="{pending_title}" onclick="rerunPending()">补跑未完成账号</button><button class="secondary" onclick="post('/production/reset-today')">管理员重置今日运行锁</button><button class="secondary" onclick="post('/stop')">停止任务</button></details></div>
<div class="card"><h3>任务状态</h3><div id="cards" class="grid"></div></div>
<div class="card"><h3>最近运行历史</h3><div id="history"></div></div>
<div class="card"><details><summary>详细日志</summary><pre id="log"></pre></details></div>
<div class="card"><div class="label">版本信息</div><div class="value">APP_VERSION {APP_VERSION}｜服务启动 {SERVICE_STARTED_AT}｜主程序修改 {datetime.fromtimestamp((BASE_DIR / "auto_all_integrated.py").stat().st_mtime).isoformat(timespec="seconds")}</div></div>
<script>
async function post(url){{let r=await fetch(url,{{method:'POST'}});alert(await r.text());loadAll();}}
async function confirmPost(url){{if(!confirm('{confirm_text}')) return; await post(url);}}
async function resumeZip(){{let p=document.getElementById('resumeZip').value; if(!p) return alert('请输入本地ZIP路径'); let r=await fetch('/resume-zip?zip_path='+encodeURIComponent(p),{{method:'POST'}}); alert(await r.text()); loadAll();}}
async function rerunPending(){{let plan=await (await fetch('/pending-accounts')).json(); if(plan.running){{return alert('\u8bf7\u5148\u505c\u6b62\u5f53\u524d\u4efb\u52a1');}} let names=plan.pending_accounts||plan.to_run||[]; if(!names.length){{return alert('\u4eca\u5929\u6ca1\u6709\u9700\u8981\u8865\u8dd1\u7684\u8d26\u53f7');}} if(names.includes('account_YOUT')){{return alert('\u5b89\u5168\u62d2\u7edd\uff1aaccount_YOUT \u5df2\u5b8c\u6210\uff0c\u4e0d\u80fd\u8fdb\u5165\u8865\u8dd1\u540d\u5355');}} let msg='\u5c06\u8865\u8dd1\u4ee5\u4e0b\u8d26\u53f7\uff1a'+names.join(', '); if((plan.failed_accounts||[]).length){{msg+='\n\u5176\u4e2d\u5931\u8d25\u8d26\u53f7\u9700\u8981\u4f60\u786e\u8ba4\uff1a'+plan.failed_accounts.join(', ');}} if(!confirm(msg+'\n\u786e\u8ba4\u7ee7\u7eed\u5417\uff1f')) return; await post('/run-pending-accounts');}}
function item(k,v){{return `<div><div class="label">${{k}}</div><div class="value">${{v??''}}</div></div>`}}
async function loadStatus(){{let r=await fetch('/status');let s=await r.json();let keys=[['当前状态','state'],['当前账号','current_account'],['当前步骤','current_step'],['已完成','completed_accounts'],['成功PDF','success_pdf_count'],['无待发订单','no_order_count'],['异常数量','exception_count'],['重复跳过','duplicate_skipped_count'],['开始时间','started_at'],['最后更新时间','updated_at']];document.getElementById('cards').innerHTML=keys.map(x=>item(x[0],s[x[1]])).join('');}}
async function loadHistory(){{let r=await fetch('/history');let h=await r.json();document.getElementById('history').innerHTML=h.map(x=>`<div>${{x.time}} | ${{x.run_id}} | ${{x.state}} | 成功 ${{x.success}} | 无订单 ${{x.no_orders}} | 异常 ${{x.exceptions}}</div>`).join('')||'暂无历史';}}
async function loadLog(){{let r=await fetch('/logs');document.getElementById('log').textContent=await r.text();}}
function loadAll(){{loadStatus();loadHistory();loadLog();}}loadAll();setInterval(loadAll,5000);
</script></main></body></html>"""
    return HTMLResponse(content=html, media_type="text/html; charset=utf-8")


@app.get("/status")
def status():
    value = current_status()
    append_history_from_status(value)
    update_latest_production_run(value)
    value["production_mode"] = sanitize_value(runtime_mode())
    value["production_can_start_today"] = can_start_production_run()
    value["app_version"] = APP_VERSION
    value["service_started_at"] = SERVICE_STARTED_AT
    return value


@app.get("/history")
def history():
    return read_json(HISTORY_PATH, [])[:10]


@app.get("/logs", response_class=PlainTextResponse)
def logs():
    return safe_log_tail()


@app.post("/run/all")
def run_all():
    if production_mode_enabled():
        return launch(["--all", "--production-run"], production=True)
    return launch(["--all"])


@app.post("/run/account/{account}")
def run_account(account: str):
    names = {a["name"] for a in read_config()["accounts"]}
    if account not in names:
        raise HTTPException(status_code=404, detail="Unknown account")
    return launch(["--account", account])


@app.post("/run/demo")
def run_demo():
    return launch(["--demo"])



@app.get("/pending-accounts")
def pending_accounts():
    return account_completion_plan()


@app.get("/run/incomplete/plan")
def run_incomplete_plan():
    return account_completion_plan()


def start_pending_accounts():
    plan = account_completion_plan()
    if plan.get("running"):
        raise HTTPException(status_code=409, detail="请先停止当前任务")
    if "account_YOUT" in plan.get("to_run", []):
        raise HTTPException(status_code=409, detail="安全拒绝：account_YOUT 已完成，不能进入补跑名单")
    if not plan["to_run"]:
        raise HTTPException(status_code=409, detail="今天没有需要补跑的账号")
    return launch(["--rerun-incomplete", "--production-run"])


@app.post("/run-pending-accounts")
def run_pending_accounts():
    return start_pending_accounts()


@app.post("/run/incomplete")
def run_incomplete():
    return start_pending_accounts()


@app.post("/resume-zip")
def resume_zip(zip_path: str, account: str = "account_YOUT", hold_sale_id: str = "", hold_error: str = "Order on hold"):
    path = Path(zip_path).expanduser()
    if not path.exists() or path.suffix.lower() != ".zip":
        raise HTTPException(status_code=404, detail="ZIP file not found")
    return launch(["--resume-zip", str(path), "--account", account, "--hold-sale-id", hold_sale_id, "--hold-error", hold_error])


@app.post("/production/reset-today")
def production_reset_today(reason: str = "manual"):
    return {"ok": True, "reset": reset_production_today(reason)}


@app.post("/stop")
def stop_task():
    global _process
    with _lock:
        if _process is None or _process.poll() is not None:
            return {"ok": True, "message": "当前没有运行中的任务"}
        write_json(STATUS_PATH, {"state": "stopping", "message": "正在停止任务", "updated_at": datetime.now().isoformat(timespec="seconds")})
        _process.terminate()
        try:
            _process.wait(timeout=10)
        except subprocess.TimeoutExpired:
            _process.kill()
            _process.wait(timeout=5)
        _process = None
        write_json(STATUS_PATH, {"state": "stopped", "message": "任务已停止", "current_step": "已停止", "updated_at": datetime.now().isoformat(timespec="seconds")})
        return {"ok": True, "message": "任务已停止"}


@app.get("/download/latest-excel")
def download_excel():
    path = latest_file("*.xlsx")
    return FileResponse(path, filename=path.name)


@app.get("/download/latest-zip")
def download_zip():
    path = latest_file("*.zip")
    return FileResponse(path, filename=path.name)


@app.get("/download/report/{filename}")
def download_report(filename: str):
    path = safe_report_path(filename, "*.xlsx")
    return FileResponse(path, filename=path.name)


@app.get("/download/archive/{filename}")
def download_archive(filename: str):
    path = safe_report_path(filename, "*.zip")
    return FileResponse(path, filename=path.name)

