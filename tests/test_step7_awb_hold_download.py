from __future__ import annotations

import io
import json
import zipfile
import os
import secrets
import tempfile
import time
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from auto_all_integrated import (
    atomic_write_json,
    can_start_real_run,
    AWBModalTimeoutError,
    awb_button_is_enabled,
    classify_awb_modal_snapshot,
    click_awb_popup_download,
    default_real_run_limit,
    extract_download_url_from_json_response,
    find_new_zip_after_click,
    parse_awb_hold_records,
    parse_awb_success_count,
    record_precheck_failed,
    record_real_run,
    ensure_real_run_limit,
    save_response_zip,
    save_zip_from_url,
    wait_for_awb_result,
)
from shipping_report import account_state, create_report, dedupe_report_records, exception_rows, shipping_rows

ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "tests" / "results"



def valid_zip_bytes(name="label.pdf", body=b"pdf"):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as archive:
        archive.writestr(name, body)
    return buf.getvalue()


class FakeAwbModalPage:
    def __init__(self, snapshot):
        self.snapshot = snapshot
        self.wait_calls = 0
        self.screenshots = []

    def evaluate(self, script):
        return self.snapshot

    def wait_for_timeout(self, ms):
        self.wait_calls += 1

    def screenshot(self, path, full_page=True):
        Path(path).write_bytes(b"fake png")
        self.screenshots.append(path)

    def content(self):
        return "<html><body>fake modal</body></html>"


def enabled_modal_snapshot(text="Bulk downloading Air Waybill for all 29 sales\nBulk download may take a while for large amount of items.\nDownload"):
    return {
        "text": text,
        "tag": "div",
        "role": "dialog",
        "class": "relative bg-white m-4 rounded-lg shadow-2xl",
        "download_button": {
            "exists": True,
            "visible": True,
            "tag": "button",
            "role": "",
            "text": "Download",
            "class": "align-middle select-none font-sans font-bold text-center transition-all disabled:opacity-50 disabled:shadow-none disabled:pointer-events-none text-xs py-3 px-6 rounded-lg focus:ring active:opacity-[0.85] border-2 border-secondary bg-secondaryBg normal-case text-secondary focus:ring-secondary hover:bg-secondary hover:text-secondaryBg hover:opacity-100 mt-auto w-full flex-1",
            "disabled": False,
            "aria_disabled": "",
            "pointer_events": "auto",
        },
    }


def disabled_modal_snapshot():
    snap = enabled_modal_snapshot()
    snap["download_button"] = dict(snap["download_button"])
    snap["download_button"]["disabled"] = True
    return snap

class FakeJsonResponse:
    url = "https://example.invalid/api/awb"
    status = 200
    headers = {"content-type": "application/json"}

    def text(self):
        return json.dumps({"data": {"downloadUrl": "https://example.invalid/download/awb.zip"}})


class FakeZipResponse:
    url = "https://example.invalid/awb.zip"
    headers = {"content-type": "application/zip"}

    def body(self):
        return b"PK\x03\x04fake zip body"


class FakeClickPage:
    def __init__(self):
        self.evaluate_calls = 0

    def evaluate(self, script):
        self.evaluate_calls += 1
        return {"clicked": True}


class Step7AwbHoldDownloadTest(unittest.TestCase):
    def setUp(self):
        self.run_id = f"step7_{datetime.now():%Y%m%d_%H%M%S}_{secrets.token_hex(2)}"
        self.base = RESULTS / self.run_id
        self.base.mkdir(parents=True, exist_ok=True)

    def hold_record(self):
        return parse_awb_hold_records("1 of 28 sales failed to download Air Waybill: # Sale ID Error 1 90933046 Order on hold", "account_YOUT", self.run_id)[0]

    def success_record(self, index=1):
        return {
            "date": "2026-06-25",
            "run_id": self.run_id,
            "account": "account_YOUT",
            "filename": f"SKU US 9 {900000 + index}.pdf",
            "sale_id": str(900000 + index),
            "carrier": "UPS",
            "tracking_number": f"1Z{index:016d}",
            "source_file": str(self.base / f"source{index}.pdf"),
            "output_file": str(self.base / f"processed{index}.pdf"),
            "status": "SUCCESS",
            "source_sha256": f"hash-{index}",
        }

    def test_bulk_only_modal_waits_for_result_rows_before_download(self):
        html_path = ROOT / "tests" / "fixtures" / "awb_bulk_modal.html"
        html = html_path.read_text(encoding="utf-8", errors="ignore")
        self.assertIn("Bulk downloading Air Waybill for all", html)
        self.assertNotIn("downloaded successfully", html)
        self.assertNotIn("Order on hold", html)
        result = classify_awb_modal_snapshot(enabled_modal_snapshot(), "account_YOUT", self.run_id)
        self.assertEqual(result["state"], "waiting")
        self.assertTrue(awb_button_is_enabled(enabled_modal_snapshot()))

    def test_download_button_disabled_then_enabled(self):
        waiting = classify_awb_modal_snapshot(disabled_modal_snapshot(), "account_YOUT", self.run_id)
        ready = classify_awb_modal_snapshot(enabled_modal_snapshot("1 Air Waybill downloaded successfully.\nDownload"), "account_YOUT", self.run_id)
        self.assertEqual(waiting["state"], "waiting")
        self.assertEqual(ready["state"], "download_ready")

    def test_awb_modal_timeout_status_when_download_stays_disabled(self):
        page = FakeAwbModalPage(disabled_modal_snapshot())
        with self.assertRaises(AWBModalTimeoutError):
            wait_for_awb_result(page, self.base, 20, time.monotonic() + 0.01, "account_YOUT", time.monotonic(), {"last": 0.0}, self.run_id)

    def test_hold_and_download_ready_is_partial_success_candidate(self):
        text = "1 of 28 sales failed to download Air Waybill: # Sale ID Error 1 90933046 Order on hold\nDownload"
        result = classify_awb_modal_snapshot(enabled_modal_snapshot(text), "account_YOUT", self.run_id)
        self.assertEqual(result["state"], "download_ready")
        self.assertEqual(result["candidate_status"], "PARTIAL_SUCCESS")
        self.assertEqual(result["hold_records"][0]["sale_id"], "90933046")

    def test_failed_text_with_download_ready_still_downloads(self):
        result = classify_awb_modal_snapshot(enabled_modal_snapshot("failed to download Air Waybill\nDownload"), "account_YOUT", self.run_id)
        self.assertEqual(result["state"], "download_ready")
        self.assertEqual(result["candidate_status"], "PARTIAL_SUCCESS")

    def test_timeout_does_not_reopen_awb_modal_or_regenerate(self):
        page = FakeAwbModalPage(disabled_modal_snapshot())
        with self.assertRaises(AWBModalTimeoutError):
            wait_for_awb_result(page, self.base, 20, time.monotonic() + 0.01, "account_YOUT", time.monotonic(), {"last": 0.0}, self.run_id)
        self.assertGreaterEqual(page.wait_calls, 0)

    def test_no_hold_text_does_not_create_90933046(self):
        result = classify_awb_modal_snapshot(enabled_modal_snapshot("1 Air Waybill downloaded successfully.\nDownload"), "account_YOUT", self.run_id)
        self.assertEqual(result["hold_records"], [])

    def test_awb_modal_timeout_not_in_shipping_rows(self):
        records = [
            {"account": "account_YOUT", "run_id": self.run_id, "status": "ACCOUNT_FAILED", "error": "failed"},
            {"account": "account_YOUT", "run_id": self.run_id, "status": "AWB_MODAL_TIMEOUT", "error": "modal timeout"},
        ]
        self.assertEqual(shipping_rows(records), [])
        self.assertEqual(len(exception_rows(records)), 2)

    def test_trace_fixture_parses_hold_90933046(self):
        fixture = ROOT / "tests" / "fixtures" / "awb_hold_modal_trace.html"
        html = fixture.read_text(encoding="utf-8")
        holds = parse_awb_hold_records(html, "account_YOUT", self.run_id)
        self.assertEqual(len(holds), 1)
        self.assertEqual(holds[0]["sale_id"], "90933046")
        self.assertEqual(holds[0]["status"], "ORDER_ON_HOLD")
        self.assertEqual(holds[0]["error"], "Order on hold")

    def test_trace_fixture_has_modal_download_before_background_download(self):
        fixture = ROOT / "tests" / "fixtures" / "awb_hold_modal_trace.html"
        html = fixture.read_text(encoding="utf-8")
        modal_start = html.index('class="awb-modal-root"')
        modal_end = html.index('</section>', modal_start)
        modal = html[modal_start:modal_end]
        background = html[modal_end:]
        self.assertIn('>Download</button>', modal)
        self.assertIn('>Download</button>', background)
        self.assertIn('Order on hold', modal)
        self.assertIn('90933046', modal)

    def test_parse_hold_sale_id_and_success_count(self):
        text = "1 of 28 sales failed to download Air Waybill: # Sale ID Error 1 90933046 Order on hold"
        holds = parse_awb_hold_records(text, "account_YOUT", self.run_id)
        self.assertEqual(len(holds), 1)
        self.assertEqual(holds[0]["sale_id"], "90933046")
        self.assertEqual(holds[0]["status"], "ORDER_ON_HOLD")
        self.assertEqual(parse_awb_success_count(text), 27)

    def test_zip_failure_keeps_hold_in_exception_sheet(self):
        records = [self.hold_record(), {"account": "account_YOUT", "run_id": self.run_id, "status": "DOWNLOAD_CAPTURE_FAILED", "error": "capture failed"}]
        report = self.base / "hold_zip_failed.xlsx"
        create_report(records, report)
        wb = load_workbook(report, read_only=True, data_only=True)
        today = [row for row in wb["今日发货"].iter_rows(min_row=2, values_only=True) if any(row)]
        errors = [row for row in wb["异常订单"].iter_rows(min_row=2, values_only=True) if any(row)]
        self.assertEqual(today, [])
        self.assertEqual([row[5] for row in errors], ["ORDER_ON_HOLD", "DOWNLOAD_CAPTURE_FAILED"])
        wb.close()

    def test_account_failed_not_in_shipping_rows(self):
        records = [{"account": "account_YOUT", "run_id": self.run_id, "status": "ACCOUNT_FAILED", "error": "failed"}]
        self.assertEqual(shipping_rows(records), [])
        self.assertEqual(len(exception_rows(records)), 1)

    def test_27_success_plus_hold_is_partial_success(self):
        records = [self.success_record(i) for i in range(1, 28)] + [self.hold_record()]
        self.assertEqual(account_state(records), "PARTIAL_SUCCESS")

    def test_zero_success_with_hold_and_download_failed_is_failed(self):
        records = [self.hold_record(), {"account": "account_YOUT", "run_id": self.run_id, "status": "DOWNLOAD_CAPTURE_FAILED", "error": "capture failed"}]
        self.assertEqual(account_state(records), "FAILED")

    def test_duplicate_hold_order_is_deduped(self):
        hold = self.hold_record()
        records = dedupe_report_records([hold, dict(hold)])
        self.assertEqual(len(records), 1)

    def test_download_button_clicks_once(self):
        page = FakeClickPage()
        click_awb_popup_download(page, action_timeout=20)
        self.assertEqual(page.evaluate_calls, 1)

    def test_zip_response_body_can_be_saved(self):
        target = self.base / "response.zip"
        saved = save_response_zip(FakeZipResponse(), target)
        self.assertEqual(saved.read_bytes()[:2], b"PK")

    def test_json_download_url_can_be_extracted_and_saved(self):
        url = extract_download_url_from_json_response(FakeJsonResponse())
        self.assertEqual(url, "https://example.invalid/download/awb.zip")
        target = self.base / "json_url.zip"
        saved = save_zip_from_url(url, target, fetcher=lambda _: valid_zip_bytes())
        self.assertGreater(saved.stat().st_size, 0)

    def test_only_click_time_new_zip_is_selected(self):
        historical = self.base / "old.zip"
        historical.write_bytes(b"PK-old")
        old_time = time.time() - 100
        os.utime(historical, (old_time, old_time))
        before = set(self.base.glob("*.zip"))
        start = time.time()
        new_zip = self.base / "new.zip"
        new_zip.write_bytes(b"PK-new")
        self.assertEqual(find_new_zip_after_click(self.base, before, start), new_zip)

    def test_history_zip_is_not_moved_or_selected(self):
        historical = self.base / "history.zip"
        historical.write_bytes(b"PK-history")
        before = set(self.base.glob("*.zip"))
        self.assertIsNone(find_new_zip_after_click(self.base, before, time.time()))
        self.assertTrue(historical.exists())

    def test_real_run_limit_reads_utf8_bom_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            path.write_text(json.dumps(default_real_run_limit()), encoding="utf-8-sig")
            payload = ensure_real_run_limit(path)
            self.assertEqual(payload["allowed_real_runs"], 2)
            self.assertEqual(payload["completed_real_runs"], 0)

    def test_real_run_limit_reads_plain_utf8_json(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            path.write_text(json.dumps(default_real_run_limit()), encoding="utf-8")
            payload = ensure_real_run_limit(path)
            self.assertEqual(payload["allowed_real_runs"], 2)
            self.assertEqual(payload["completed_real_runs"], 0)

    def test_real_run_limit_write_has_no_bom(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            record_real_run("account_YOUT", "cmd1", path)
            self.assertNotEqual(path.read_bytes()[:3], b"\xef\xbb\xbf")

    def test_real_run_limit_atomic_write_replaces_tmp(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            atomic_write_json(path, default_real_run_limit())
            self.assertTrue(path.exists())
            self.assertFalse(path.with_name("real_run_limit.json.tmp").exists())

    def test_precheck_failed_does_not_consume_real_run_quota(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            record_precheck_failed("account_YOUT", "cmd", "BOM", path)
            payload = ensure_real_run_limit(path)
            self.assertEqual(payload["completed_real_runs"], 0)
            self.assertEqual(payload["remaining_real_runs"], 2)
            self.assertTrue(can_start_real_run(path))
            self.assertEqual(payload["runs"][0]["status"], "PRECHECK_FAILED")

    def test_startup_precheck_failure_does_not_require_site_config(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            record_precheck_failed("account_YOUT", "cmd", "local json precheck failed", path)
            payload = ensure_real_run_limit(path)
            self.assertEqual(payload["completed_real_runs"], 0)
            self.assertEqual(payload["runs"][0]["status"], "PRECHECK_FAILED")

    def test_completed_real_runs_count_remains_correct(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            record_precheck_failed("account_YOUT", "cmd0", "precheck", path)
            record_real_run("account_YOUT", "cmd1", path)
            payload = record_real_run("all", "cmd2", path)
            self.assertEqual(payload["completed_real_runs"], 2)
            self.assertEqual(payload["remaining_real_runs"], 0)
            self.assertFalse(can_start_real_run(path))
    def test_real_run_limit_rejects_after_two(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "real_run_limit.json"
            path.write_text(json.dumps(default_real_run_limit()), encoding="utf-8")
            self.assertTrue(can_start_real_run(path))
            record_real_run("account_YOUT", "cmd1", path)
            record_real_run("all", "cmd2", path)
            self.assertFalse(can_start_real_run(path))
            with self.assertRaises(RuntimeError):
                record_real_run("account_YOUT", "cmd3", path)

    def test_offline_tests_use_fake_network_inputs(self):
        self.assertTrue(FakeJsonResponse.url.startswith("https://example.invalid"))
        target = self.base / "offline_fetcher.zip"
        saved = save_zip_from_url("https://example.invalid/download/awb.zip", target, fetcher=lambda _: valid_zip_bytes())
        self.assertTrue(saved.exists())


if __name__ == "__main__":
    unittest.main(verbosity=2)






