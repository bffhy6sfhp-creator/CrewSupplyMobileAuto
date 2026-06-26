from __future__ import annotations

import secrets
import shutil
import unittest
from datetime import datetime, timedelta
from pathlib import Path

try:
    import pymupdf as fitz
except ImportError:
    import fitz
from openpyxl import load_workbook

from pdf_processor import collect_input_pdfs, process_pdf_list, sha256_file
from shipping_report import create_report

ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "tests" / "results"
ADDRESS = ROOT / "assets" / "address.png"


def make_pdf(path: Path, lines: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    y = 40
    for line in lines:
        page.insert_text((30, y), line, fontsize=14)
        y += 30
    doc.save(path)
    doc.close()


class Step3DedupPipelineTest(unittest.TestCase):
    def setUp(self):
        self.run_id = f"step3_{datetime.now():%Y%m%d_%H%M%S}_{secrets.token_hex(2)}"
        self.base = RESULTS / self.run_id
        self.input = self.base / "input" / datetime.now().strftime("%Y-%m-%d") / "Original"
        self.processed = self.base / "Processed"
        self.unknown = self.base / "Unknown"
        self.failed = self.base / "Failed"
        self.base.mkdir(parents=True, exist_ok=True)

    def test_same_pdf_copied_twice_outputs_once(self):
        first = self.input / "SKU US 9 900001.pdf"
        second = self.input / "SKU US 9 900001-copy.pdf"
        make_pdf(first, ["FedEx", "TRK# 1234 5678 9012", "INTL ECONOMY", "ORIGIN ID", "BILL SENDER"])
        shutil.copy2(first, second)
        records = process_pdf_list([first, second], self.processed, self.unknown, self.failed, ADDRESS, "acct", self.run_id)
        self.assertEqual(sum(r["status"] == "SUCCESS" for r in records), 1)
        self.assertEqual(sum(r["status"] == "DUPLICATE_SKIPPED" for r in records), 1)
        self.assertEqual(len(list(self.processed.glob("*.pdf"))), 1)

    def test_same_filename_different_content_outputs_two(self):
        one = self.input / "A" / "SKU US 10 900002.pdf"
        two = self.input / "B" / "SKU US 10 900002.pdf"
        make_pdf(one, ["FedEx", "TRK# 1234 5678 9012", "INTL ECONOMY", "ORIGIN ID", "BILL SENDER"])
        make_pdf(two, ["FedEx", "TRK# 9999 8888 7777", "INTL ECONOMY", "ORIGIN ID", "BILL SENDER"])
        self.assertNotEqual(sha256_file(one), sha256_file(two))
        records = process_pdf_list([one, two], self.processed, self.unknown, self.failed, ADDRESS, "acct", self.run_id)
        self.assertEqual(sum(r["status"] == "SUCCESS" for r in records), 2)
        outputs = [Path(r["output_file"]).name for r in records if r["status"] == "SUCCESS"]
        self.assertEqual(len(outputs), 2)
        self.assertEqual(len(set(outputs)), 2)

    def test_processed_dir_not_scanned_and_yesterday_skipped(self):
        today = datetime.now().strftime("%Y-%m-%d")
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        valid = self.base / "Label" / "acct" / today / "Original" / "valid.pdf"
        processed = self.base / "Label" / "acct" / today / "Processed" / "old.pdf"
        old = self.base / "Label" / "acct" / yesterday / "Original" / "old.pdf"
        for path in [valid, processed, old]:
            make_pdf(path, ["FedEx", "TRK# 1234 5678 9012", "INTL ECONOMY", "ORIGIN ID", "BILL SENDER"])
        found = collect_input_pdfs(self.base / "Label", run_date=today)
        self.assertEqual(found, [valid])

    def test_excel_does_not_write_duplicate_shipping_rows(self):
        source = self.input / "SKU US 11 900003.pdf"
        make_pdf(source, ["FedEx", "TRK# 1234 5678 9012", "INTL ECONOMY", "ORIGIN ID", "BILL SENDER"])
        records = process_pdf_list([source, source], self.processed, self.unknown, self.failed, ADDRESS, "acct", self.run_id)
        report = self.base / "report.xlsx"
        create_report(records, report)
        wb = load_workbook(report, read_only=True, data_only=True)
        ws = wb["今日发货"]
        data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
        wb.close()
        self.assertEqual(len(data_rows), 1)


if __name__ == "__main__":
    unittest.main(verbosity=2)



