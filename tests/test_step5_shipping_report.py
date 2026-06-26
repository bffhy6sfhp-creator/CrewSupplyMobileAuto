from __future__ import annotations

import secrets
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from shipping_report import create_report

ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "tests" / "results"


class Step5ShippingReportTest(unittest.TestCase):
    def setUp(self):
        self.run_id = f"step5_{datetime.now():%Y%m%d_%H%M%S}_{secrets.token_hex(2)}"
        self.base = RESULTS / self.run_id
        self.base.mkdir(parents=True, exist_ok=True)
        self.report = self.base / "shipping_report.xlsx"

    def make_records(self):
        success = {
            "date": "2026-06-25", "run_id": self.run_id, "account": "account_YOUT",
            "filename": "SKU US 9 900001.pdf", "sale_id": "900001", "sku": "SKU", "size": "US 9", "quantity": "1",
            "carrier": "UPS", "tracking_number": "1Z999AA10123456784",
            "source_file": str(ROOT / "Label" / "x" / "source.pdf"),
            "output_file": str(ROOT / "Label" / "x" / "processed.pdf"),
            "status": "SUCCESS", "message": "ok", "source_sha256": "hash-success",
        }
        duplicate_same = dict(success)
        duplicate_same["filename"] = "SKU US 9 900001 copy.pdf"
        duplicate_same["status"] = "SUCCESS"
        skipped = dict(success)
        skipped["status"] = "DUPLICATE_SKIPPED"
        skipped["source_sha256"] = "hash-success"
        failed = {
            "date": "2026-06-25", "run_id": self.run_id, "account": "account_YOUT",
            "filename": "BAD 900004.pdf", "sale_id": "900004", "carrier": "UNKNOWN",
            "source_file": str(ROOT / "Label" / "bad.pdf"), "status": "PDF_PROCESS_FAILED",
            "message": "full stack hidden", "error": "PDF crashed with a very long stack trace " + "x" * 300,
            "source_sha256": "hash-failed",
        }
        return [
            success, duplicate_same, skipped, failed,
            {"date": "2026-06-25", "run_id": self.run_id, "account": "account_KAKOMAY", "status": "NO_ORDERS", "message": "无待发订单"},
            {"date": "2026-06-25", "run_id": self.run_id, "account": "account_SUP", "status": "NO_ORDERS", "message": "无待发订单"},
        ]

    def test_report_rules_and_summary(self):
        create_report(self.make_records(), self.report)
        wb = load_workbook(self.report, read_only=False, data_only=True)
        self.assertEqual(wb.sheetnames, ["今日发货", "异常订单", "账号汇总"])
        self.assertFalse(wb["今日发货"].merged_cells.ranges)
        self.assertEqual(wb["今日发货"].freeze_panes, "A2")
        self.assertIsNotNone(wb["今日发货"].auto_filter.ref)

        today_rows = [row for row in wb["今日发货"].iter_rows(min_row=2, values_only=True) if any(row)]
        self.assertEqual(len(today_rows), 1)  # only successful shipping PDFs enter 今日发货
        statuses = [row[11] for row in today_rows]
        self.assertEqual(statuses.count("SUCCESS"), 1)
        self.assertEqual(statuses.count("PDF_PROCESS_FAILED"), 0)

        error_rows = [row for row in wb["异常订单"].iter_rows(min_row=2, values_only=True) if any(row)]
        self.assertEqual(len(error_rows), 1)
        self.assertEqual(error_rows[0][5], "PDF_PROCESS_FAILED")
        self.assertLessEqual(len(error_rows[0][6]), 180)
        self.assertNotIn("NO_ORDERS", [row[5] for row in error_rows])

        summary_rows = {row[0]: row for row in wb["账号汇总"].iter_rows(min_row=2, values_only=True) if row[0]}
        self.assertEqual(summary_rows["account_KAKOMAY"][1], "NO_ORDERS")
        self.assertEqual(summary_rows["account_SUP"][1], "NO_ORDERS")
        self.assertEqual(summary_rows["account_YOUT"][1], "PARTIAL_SUCCESS")
        self.assertEqual(summary_rows["account_YOUT"][3], 1)
        self.assertEqual(summary_rows["account_YOUT"][8], 1)
        self.assertEqual(summary_rows["account_YOUT"][9], 1)
        wb.close()


if __name__ == "__main__":
    unittest.main(verbosity=2)

