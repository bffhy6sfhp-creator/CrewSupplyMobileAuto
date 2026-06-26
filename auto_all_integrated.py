from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import secrets
import shutil
import sys
import time
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.request import urlopen

from pdf_processor import process_folder, process_pdf_list
from shipping_report import create_report, zip_processed

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
STATUS_PATH = BASE_DIR / "status.json"
LABEL_DIR = BASE_DIR / "Label"
REPORT_DIR = BASE_DIR / "Reports"
ADDRESS_FILE = BASE_DIR / "assets" / "address.png"
REAL_RUN_LIMIT_PATH = BASE_DIR / "data" / "real_run_limit.json"
APP_VERSION = "2.0.0-stable"

NO_ORDER_PATTERNS = [
    r"\b0\s+(?:to\s+ship\s+)?(?:orders?|sales?|items?|records?|results?)\b",
    r"\bno\s+(?:to\s+ship\s+)?(?:orders?|sales?|items?|records?|results?|data)\b",
    r"no matching records",
    r"nothing to display",
    r"there are no .*?(?:orders|sales|items|records)",
    r"无待发订单",
    r"没有待发订单",
    r"暂无数据",
]
DISABLED_CLASS_MARKERS = ["pointer-events-none", "cursor-not-allowed", "opacity-50"]


class AccountTimeoutError(TimeoutError):
    pass


class DownloadCaptureFailed(RuntimeError):
    pass


class AWBModalTimeoutError(TimeoutError):
    pass


def log(message: str) -> None:
    print(message, flush=True)


def generate_run_id() -> str:
    return f"{datetime.now():%Y%m%d_%H%M%S}_{secrets.token_hex(2)}"


def load_config() -> Dict:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8-sig"))


def _atomic_write_status(payload: Dict) -> None:
    STATUS_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = STATUS_PATH.with_name(STATUS_PATH.name + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(STATUS_PATH)


def read_status() -> Dict:
    try:
        return json.loads(STATUS_PATH.read_text(encoding="utf-8-sig")) if STATUS_PATH.exists() else {}
    except Exception:
        return {}


def write_status(state: str, message: str, **extra) -> None:
    payload = {
        "state": state,
        "message": message,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "pid": os.getpid(),
        "app_version": APP_VERSION,
        **extra,
    }
    _atomic_write_status(payload)
    log(f"[{state}] {message}")


def update_status(**extra) -> None:
    payload = read_status()
    payload.update(extra)
    payload["updated_at"] = datetime.now().isoformat(timespec="seconds")
    payload["pid"] = os.getpid()
    payload["app_version"] = APP_VERSION
    _atomic_write_status(payload)


def normalize_us_size(size: str) -> str:
    raw = str(size).strip().upper().replace("US", "").replace(" ", "")
    if not raw:
        return str(size)
    if "." in raw:
        return raw.rstrip("0").rstrip(".")
    if not raw.isdigit():
        return str(size)
    number = int(raw)
    if number >= 30 and number % 10 == 0:
        return str(number // 10)
    if number >= 30:
        return f"{number // 10}.{number % 10}"
    return raw


def make_new_name(filename: str) -> str:
    stem, ext = Path(filename).stem, Path(filename).suffix
    parts = stem.split("_")
    if len(parts) < 2:
        return Path(filename).name
    sale_id, last = parts[0], parts[-1]
    match = re.match(r"(.+)-([A-Za-z0-9.]+)$", last)
    if not match:
        return Path(filename).name
    model, size = match.groups()
    final_size = f"US {normalize_us_size(size)}" if size.replace(".", "").isdigit() else size.upper()
    return f"{model} {final_size} {sale_id}{ext}"


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    suffix = secrets.token_hex(2)
    candidate = path.with_name(f"{path.stem}__{suffix}{path.suffix}")
    if not candidate.exists():
        return candidate
    index = 2
    while True:
        candidate = path.with_name(f"{path.stem}__{suffix}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def account_run_dirs(account: str, run_id: str | None = None) -> Dict[str, Path]:
    day_root = LABEL_DIR / account / datetime.now().strftime("%Y-%m-%d")
    suffix = Path(run_id) if run_id else Path()
    dirs = {
        "root": day_root,
        "downloads": day_root / "Downloads" / suffix,
        "original": day_root / "Original" / suffix,
        "processed": day_root / "Processed" / suffix,
        "unknown": day_root / "Unknown" / suffix,
        "failed": day_root / "Failed" / suffix,
        "screenshots": day_root / "Screenshots" / suffix,
    }
    for path in dirs.values():
        path.mkdir(parents=True, exist_ok=True)
    return dirs


def save_page_artifacts(page, folder: Path, reason: str) -> None:
    safe = re.sub(r"[^A-Za-z0-9_-]", "_", reason)[:50]
    stamp = datetime.now().strftime("%H%M%S")
    try:
        screenshot = unique_path(folder / f"{stamp}_{safe}.png")
        page.screenshot(path=str(screenshot), full_page=True)
        log(f"Screenshot: {screenshot}")
    except Exception as exc:
        log(f"Screenshot failed: {exc}")
    try:
        html_path = unique_path(folder / f"{stamp}_{safe}.html")
        html_path.write_text(page.content(), encoding="utf-8", errors="ignore")
        log(f"HTML summary: {html_path}")
    except Exception as exc:
        log(f"HTML summary failed: {exc}")


def is_page_failure(exc: Exception) -> bool:
    message = str(exc).lower()
    return any(token in message for token in [
        "timeout", "timed out", "target closed", "page closed", "browser closed", "context closed",
    ])


def get_page_text(page) -> str:
    try:
        return page.locator("body").inner_text(timeout=3000)
    except Exception as exc:
        if is_page_failure(exc):
            raise
        return ""


def looks_like_no_orders(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", (text or "").lower()).strip()
    return any(re.search(pattern, normalized, re.I) for pattern in NO_ORDER_PATTERNS)


def no_orders_record(account: str, run_id: str, message: str = "无待发订单") -> Dict:
    return {
        "account": account,
        "run_id": run_id,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "filename": "",
        "source_file": "",
        "processed_file": "",
        "output_file": "",
        "carrier": "",
        "tracking": "",
        "tracking_number": "",
        "status": "NO_ORDERS",
        "action": message,
        "message": message,
        "error": "",
    }


def error_record(account: str, run_id: str, status: str, message: str) -> Dict:
    return {
        "account": account,
        "run_id": run_id,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "filename": "",
        "source_file": "",
        "processed_file": "",
        "output_file": "",
        "carrier": "",
        "tracking": "",
        "tracking_number": "",
        "status": status,
        "action": message,
        "message": message,
        "error": message,
    }



def order_on_hold_record(account: str, run_id: str, sale_id: str, reason: str = "Order on hold") -> Dict:
    record = error_record(account, run_id, "ORDER_ON_HOLD", reason)
    record["sale_id"] = str(sale_id or "").strip()
    record["filename"] = f"{record['sale_id']}.pdf" if record["sale_id"] else ""
    return record


def parse_awb_success_count(text: str) -> int:
    normalized = re.sub(r"\s+", " ", text or " ").strip()
    match = re.search(r"(\d+)\s+Air\s+Waybill\s+downloaded\s+successfully", normalized, re.I)
    if match:
        return int(match.group(1))
    partial = re.search(r"(\d+)\s+of\s+(\d+)\s+sales?\s+failed\s+to\s+download\s+Air\s+Waybill", normalized, re.I)
    if partial:
        failed, total = int(partial.group(1)), int(partial.group(2))
        return max(total - failed, 0)
    return 0


def parse_awb_hold_records(text: str, account: str, run_id: str) -> List[Dict]:
    records: List[Dict] = []
    seen = set()
    normalized = re.sub(r"\s+", " ", text or " ").strip()
    patterns = [
        r"(?P<sale>\d{6,})\s+(?P<reason>Order\s+on\s+hold)",
        r"(?P<sale>\d{6,}).{0,80}?(?P<reason>Order\s+on\s+hold)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, normalized, re.I):
            sale_id = match.group("sale")
            reason = re.sub(r"\s+", " ", match.group("reason")).strip()
            key = (sale_id, reason.lower())
            if key in seen:
                continue
            seen.add(key)
            records.append(order_on_hold_record(account, run_id, sale_id, reason))
    return records


def default_real_run_limit() -> Dict:
    return {
        "allowed_real_runs": 2,
        "completed_real_runs": 0,
        "remaining_real_runs": 2,
        "runs": [],
        "policy": {
            "1": "account_YOUT only after explicit confirmation",
            "2": "all three accounts final validation after explicit confirmation",
        },
    }


def normalize_real_run_limit(payload: Dict) -> Dict:
    allowed = int(payload.get("allowed_real_runs", payload.get("max_real_runs", 2)))
    runs = list(payload.get("runs", []))
    completed = sum(1 for run in runs if run.get("status") != "PRECHECK_FAILED")
    if "completed_real_runs" in payload:
        completed = int(payload.get("completed_real_runs") or 0)
    elif "used_real_runs" in payload and not runs:
        completed = int(payload.get("used_real_runs") or 0)
    payload["allowed_real_runs"] = allowed
    payload["completed_real_runs"] = completed
    payload["remaining_real_runs"] = max(allowed - completed, 0)
    payload["runs"] = runs
    payload.setdefault("policy", default_real_run_limit()["policy"])
    payload.pop("max_real_runs", None)
    payload.pop("used_real_runs", None)
    return payload


def atomic_write_json(path: Path, payload: Dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f"{path.name}.tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def ensure_real_run_limit(path: Path = REAL_RUN_LIMIT_PATH) -> Dict:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        payload = default_real_run_limit()
        atomic_write_json(path, payload)
        return payload
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    normalized = normalize_real_run_limit(payload)
    return normalized


def save_real_run_limit(payload: Dict, path: Path = REAL_RUN_LIMIT_PATH) -> Dict:
    normalized = normalize_real_run_limit(payload)
    atomic_write_json(path, normalized)
    return normalized


def can_start_real_run(path: Path = REAL_RUN_LIMIT_PATH) -> bool:
    payload = ensure_real_run_limit(path)
    return int(payload.get("completed_real_runs", 0)) < int(payload.get("allowed_real_runs", 2))


def record_real_run(account_scope: str, command: str, path: Path = REAL_RUN_LIMIT_PATH, status: str = "COMPLETED") -> Dict:
    payload = ensure_real_run_limit(path)
    completed = int(payload.get("completed_real_runs", 0))
    allowed = int(payload.get("allowed_real_runs", 2))
    if status != "PRECHECK_FAILED" and completed >= allowed:
        raise RuntimeError("REAL_RUN_LIMIT_REACHED")
    run = {
        "time": datetime.now().isoformat(timespec="seconds"),
        "account_scope": account_scope,
        "command": command,
        "status": status,
    }
    payload.setdefault("runs", []).append(run)
    if status != "PRECHECK_FAILED":
        completed += 1
    payload["completed_real_runs"] = completed
    payload["remaining_real_runs"] = max(allowed - completed, 0)
    return save_real_run_limit(payload, path)


def record_precheck_failed(account_scope: str, command: str, message: str, path: Path = REAL_RUN_LIMIT_PATH) -> Dict:
    payload = ensure_real_run_limit(path)
    payload.setdefault("runs", []).append({
        "time": datetime.now().isoformat(timespec="seconds"),
        "account_scope": account_scope,
        "command": command,
        "status": "PRECHECK_FAILED",
        "message": message,
    })
    return save_real_run_limit(payload, path)


def diagnostic_path(dirs: Dict[str, Path], account: str) -> Path:
    return unique_path(dirs["screenshots"] / f"{account}_download_diagnostics.json")


def write_download_diagnostics(diagnostics: Dict, dirs: Dict[str, Path], account: str) -> Path:
    path = diagnostic_path(dirs, account)
    safe = dict(diagnostics)
    safe["written_at"] = datetime.now().isoformat(timespec="seconds")
    path.write_text(json.dumps(safe, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def save_zip_bytes_dedup(body: bytes, target: Path) -> Path:
    if not body.startswith(b"PK"):
        raise RuntimeError("Downloaded body is not a ZIP")
    target.parent.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(body).hexdigest()
    for existing in target.parent.glob("*.zip"):
        try:
            if existing.exists() and existing.stat().st_size > 0 and sha256_file(existing) == digest:
                return existing
        except OSError:
            continue
    target.write_bytes(body)
    with zipfile.ZipFile(target) as archive:
        bad = archive.testzip()
        if bad:
            raise RuntimeError(f"ZIP integrity check failed at: {bad}")
    return target


def find_new_zip_after_click(download_dir: Path, before_zips, start_time: float) -> Path | None:
    candidates = []
    for path in download_dir.glob("*.zip"):
        try:
            if path in before_zips:
                continue
            stat = path.stat()
            if stat.st_mtime + 1 < start_time or stat.st_size <= 0:
                continue
            candidates.append(path)
        except OSError:
            continue
    if not candidates:
        return None
    return max(candidates, key=lambda item: item.stat().st_mtime)


def extract_download_url_from_json_payload(payload) -> str:
    if isinstance(payload, dict):
        for value in payload.values():
            found = extract_download_url_from_json_payload(value)
            if found:
                return found
    elif isinstance(payload, list):
        for value in payload:
            found = extract_download_url_from_json_payload(value)
            if found:
                return found
    elif isinstance(payload, str):
        if re.search(r"https?://", payload, re.I) and re.search(r"(zip|download|air.?waybill|awb)", payload, re.I):
            return payload
    return ""


def extract_download_url_from_json_response(response) -> str:
    try:
        headers = {k.lower(): v.lower() for k, v in response.headers.items()}
        if "json" not in headers.get("content-type", ""):
            return ""
        payload = json.loads(response.text())
        return extract_download_url_from_json_payload(payload)
    except Exception:
        return ""


def save_zip_from_url(url: str, target: Path, fetcher=None) -> Path:
    fetch = fetcher or (lambda value: urlopen(value, timeout=20).read())
    body = fetch(url)
    return save_zip_bytes_dedup(body, target)
def locator_has_disabled_state(locator) -> bool:
    try:
        if not locator.is_enabled():
            return True
    except Exception:
        pass
    for attr in ["disabled", "aria-disabled"]:
        try:
            value = locator.get_attribute(attr)
        except Exception:
            value = None
        if attr == "disabled" and value is not None:
            return True
        if attr == "aria-disabled" and str(value).lower() == "true":
            return True
    try:
        class_name = locator.get_attribute("class") or ""
    except Exception:
        class_name = ""
    return any(marker in class_name for marker in DISABLED_CLASS_MARKERS)


def find_visible_text_locator(page, text: str):
    locator = page.get_by_text(text, exact=True)
    try:
        count = locator.count()
        for index in range(count):
            candidate = locator.nth(index)
            if candidate.is_visible():
                return candidate
    except Exception as exc:
        if is_page_failure(exc):
            raise
    return None


def has_selectable_orders(page) -> bool:
    for selector in ["input[type='checkbox']", "[role='checkbox']"]:
        boxes = page.locator(selector)
        try:
            count = boxes.count()
        except Exception as exc:
            if is_page_failure(exc):
                raise
            continue
        for index in range(count):
            box = boxes.nth(index)
            try:
                if box.is_visible() and not locator_has_disabled_state(box):
                    return True
            except Exception as exc:
                if is_page_failure(exc):
                    raise
                continue
    return False


def evaluate_to_ship_download_state(page) -> str:
    if looks_like_no_orders(get_page_text(page)):
        return "no_orders"
    if not has_selectable_orders(page):
        return "no_orders"
    button = find_visible_text_locator(page, "Download All")
    if button is None or locator_has_disabled_state(button):
        return "no_orders"
    return "ready"


def check_deadline(deadline: float, account: str, step: str) -> None:
    if time.monotonic() > deadline:
        raise AccountTimeoutError(f"{account} exceeded account timeout during {step}")


def heartbeat(account: str, step: str, start: float, state: Dict[str, float], force: bool = False) -> None:
    now = time.monotonic()
    if force or now - state.get("last", 0) >= 20:
        elapsed = int(now - start)
        log(f"[heartbeat] 当前账号={account} 当前步骤={step} 已运行秒数={elapsed}")
        state["last"] = now


def stage(account: str, step: str, start: float, heartbeat_state: Dict[str, float]) -> None:
    log(f"[{account}] {step}")
    update_status(state="running", message=f"Processing {account}", current_account=account, current_step=step)
    heartbeat(account, step, start, heartbeat_state, force=True)



def find_awb_dialog(page):
    dialogs = page.locator("[role='dialog']")
    try:
        count = dialogs.count()
        for index in range(count - 1, -1, -1):
            candidate = dialogs.nth(index)
            if candidate.is_visible() and "AIR WAYBILL" in candidate.inner_text(timeout=1000).upper():
                return candidate
    except Exception:
        pass
    return page.locator("body")




def awb_button_is_enabled(snapshot: Dict) -> bool:
    button = snapshot.get("download_button") or {}
    if not button.get("exists") or not button.get("visible", True):
        return False
    if button.get("disabled"):
        return False
    if str(button.get("aria_disabled", "")).lower() == "true":
        return False
    if str(button.get("pointer_events", "")).lower() == "none":
        return False
    class_tokens = set(str(button.get("class", "") or "").split())
    return not any(marker in class_tokens for marker in ["pointer-events-none", "cursor-not-allowed"])


def classify_awb_modal_snapshot(snapshot: Dict, account: str, run_id: str) -> Dict:
    text = snapshot.get("text", "") or ""
    normalized = re.sub(r"\s+", " ", text)
    hold_records = parse_awb_hold_records(text, account, run_id)
    success_count = parse_awb_success_count(text)
    button_ready = awb_button_is_enabled(snapshot)
    has_failed = "failed" in normalized.lower()
    has_result_text = bool(re.search(r"Sale ID|Order on hold|failed to download Air Waybill|downloaded successfully", normalized, re.I))
    bulk_only = bool(re.search(r"Bulk downloading Air Waybill", normalized, re.I)) and not has_result_text
    if button_ready and not bulk_only:
        candidate = "PARTIAL_SUCCESS" if hold_records or has_failed else "SUCCESS"
        return {
            "state": "download_ready",
            "candidate_status": candidate,
            "success_count": success_count,
            "hold_records": hold_records,
            "text": text,
            "button": snapshot.get("download_button") or {},
        }
    if has_failed and not button_ready:
        return {
            "state": "failed_no_download",
            "candidate_status": "FAILED",
            "success_count": success_count,
            "hold_records": hold_records,
            "text": text,
            "button": snapshot.get("download_button") or {},
        }
    return {
        "state": "waiting",
        "candidate_status": "WAITING",
        "success_count": success_count,
        "hold_records": hold_records,
        "text": text,
        "button": snapshot.get("download_button") or {},
    }
def get_awb_modal_snapshot(page) -> Dict:
    script = r"""
    () => {
      const visible = (el) => {
        if (!el) return false;
        const style = window.getComputedStyle(el);
        const rect = el.getBoundingClientRect();
        return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
      };
      const textOf = (el) => (el && (el.innerText || el.textContent || el.value || '')) || '';
      const exactDownload = (el) => /^\s*Download\s*$/.test(textOf(el));
      const clickableSelector = 'button,[role="button"],a,input[type="button"],input[type="submit"],[onclick]';
      const clickableDownload = (root) => Array.from(root.querySelectorAll(clickableSelector)).find(exactDownload);
      const hasResultText = (el) => /Sale ID|Air Waybill downloaded|Order on hold|failed to download Air Waybill|downloaded successfully/i.test(textOf(el));
      const titles = Array.from(document.querySelectorAll('h1,h2,h3,h4,h5,h6,div,span')).filter(el => visible(el) && /AWB Download/i.test(textOf(el)));
      let container = null;
      for (const title of titles.reverse()) {
        let cur = title;
        while (cur && cur !== document.body) {
          const text = textOf(cur);
          if (/AWB Download/i.test(text) && hasResultText(cur) && clickableDownload(cur)) {
            container = cur;
            break;
          }
          cur = cur.parentElement;
        }
        if (container) break;
      }
      if (!container) {
        const fallback = Array.from(document.querySelectorAll('[data-floating-ui-portal] > div, .fixed, .relative, [role="dialog"]'));
        container = fallback.reverse().find(el => visible(el) && /AWB Download|Bulk downloading Air Waybill|Air Waybill downloaded|Order on hold|failed to download Air Waybill/i.test(textOf(el)) && clickableDownload(el));
      }
      if (!container) container = document.body;
      const button = clickableDownload(container);
      const rect = button ? button.getBoundingClientRect() : null;
      const style = button ? window.getComputedStyle(button) : null;
      return {
        text: textOf(container),
        tag: container.tagName ? container.tagName.toLowerCase() : '',
        role: container.getAttribute ? (container.getAttribute('role') || '') : '',
        class: container.getAttribute ? (container.getAttribute('class') || '') : '',
        download_button: button ? {
          exists: true,
          visible: visible(button),
          tag: button.tagName.toLowerCase(),
          role: button.getAttribute('role') || '',
          text: textOf(button).trim(),
          class: button.getAttribute('class') || '',
          disabled: button.hasAttribute('disabled') || button.disabled === true,
          aria_disabled: button.getAttribute('aria-disabled') || '',
          pointer_events: style ? style.pointerEvents : '',
          visibility: style ? style.visibility : '',
          bounding_box: rect ? {x: rect.x, y: rect.y, width: rect.width, height: rect.height} : null
        } : {exists: false}
      };
    }
    """
    try:
        return page.evaluate(script) or {"text": "", "download_button": {"exists": False}}
    except Exception:
        return {"text": get_page_text(page), "download_button": {"exists": False}}


def get_awb_dialog_text(page) -> str:
    return get_awb_modal_snapshot(page).get("text", "")
def click_awb_popup_download(page, action_timeout: int) -> None:
    script = r"""
    () => {
      const visible = (el) => {
        if (!el) return false;
        const style = window.getComputedStyle(el);
        const rect = el.getBoundingClientRect();
        return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
      };
      const textOf = (el) => (el && (el.innerText || el.textContent || el.value || '')) || '';
      const exactDownload = (el) => /^\s*Download\s*$/.test(textOf(el));
      const clickableSelector = 'button,[role="button"],a,input[type="button"],input[type="submit"],[onclick]';
      const clickableDownload = (root) => Array.from(root.querySelectorAll(clickableSelector)).find(exactDownload);
      const hasResultText = (el) => /Sale ID|Air Waybill downloaded|Order on hold|failed to download Air Waybill|downloaded successfully/i.test(textOf(el));
      const titles = Array.from(document.querySelectorAll('h1,h2,h3,h4,h5,h6,div,span')).filter(el => visible(el) && /AWB Download/i.test(textOf(el)));
      let container = null;
      for (const title of titles.reverse()) {
        let cur = title;
        while (cur && cur !== document.body) {
          const text = textOf(cur);
          if (/AWB Download/i.test(text) && hasResultText(cur) && clickableDownload(cur)) {
            container = cur;
            break;
          }
          cur = cur.parentElement;
        }
        if (container) break;
      }
      if (!container) {
        const fallback = Array.from(document.querySelectorAll('[data-floating-ui-portal] > div, .fixed, .relative, [role="dialog"]'));
        container = fallback.reverse().find(el => visible(el) && /AWB Download|Bulk downloading Air Waybill|Air Waybill downloaded|Order on hold|failed to download Air Waybill/i.test(textOf(el)) && clickableDownload(el));
      }
      if (!container) return {clicked:false, reason:'AWB modal container not found'};
      const button = clickableDownload(container);
      if (!button) return {clicked:false, reason:'Download button not found in AWB modal', text: textOf(container)};
      const style = window.getComputedStyle(button);
      const rect = button.getBoundingClientRect();
      const classTokens = new Set((button.getAttribute('class') || '').split(/\s+/).filter(Boolean));
      const blocked = !visible(button) || button.hasAttribute('disabled') || button.disabled === true || button.getAttribute('aria-disabled') === 'true' || style.pointerEvents === 'none' || classTokens.has('pointer-events-none') || classTokens.has('cursor-not-allowed');
      const meta = {
        tag: button.tagName.toLowerCase(),
        role: button.getAttribute('role') || '',
        text: textOf(button).trim(),
        class: button.getAttribute('class') || '',
        disabled: button.hasAttribute('disabled') || button.disabled === true,
        aria_disabled: button.getAttribute('aria-disabled') || '',
        pointer_events: style.pointerEvents,
        visibility: style.visibility,
        bounding_box: {x: rect.x, y: rect.y, width: rect.width, height: rect.height}
      };
      if (blocked) return {clicked:false, reason:'Download button disabled or not actionable', button: meta, text: textOf(container)};
      button.scrollIntoView({block:'center', inline:'center'});
      button.click();
      return {clicked:true, text: textOf(button), button: meta};
    }
    """
    result = page.evaluate(script)
    if result and result.get("clicked"):
        log("[download] 点击弹窗Download")
        return
    raise TimeoutError(f"AWB popup Download button was not clickable: {result}")
def wait_for_awb_result(page, screenshot_dir: Path, action_timeout: int, deadline: float, account: str, start: float, hb: Dict[str, float], run_id: str) -> Dict:
    end_time = min(time.monotonic() + 60, deadline)
    last_snapshot: Dict = {"text": "", "download_button": {"exists": False}}
    consecutive_ready = 0
    while time.monotonic() < end_time:
        heartbeat(account, "生成AWB", start, hb)
        snapshot = get_awb_modal_snapshot(page)
        last_snapshot = snapshot
        result = classify_awb_modal_snapshot(snapshot, account, run_id)
        button_ready = awb_button_is_enabled(snapshot)
        consecutive_ready = consecutive_ready + 1 if button_ready else 0
        summary = re.sub(r"\s+", " ", result.get("text", ""))[:220]
        button = result.get("button", {})
        log(f"[awb] modal_state={result['state']} button_exists={button.get('exists')} enabled={button_ready} text={summary}")
        if result["state"] == "download_ready":
            if result.get("hold_records"):
                save_page_artifacts(page, screenshot_dir, "awb_hold_rows")
            return result
        # Some CrewSupply responses leave only the bulk message while the real Download
        # button is already actionable. Two stable snapshots avoid clicking a transient UI.
        if consecutive_ready >= 2:
            fallback = dict(result)
            fallback["state"] = "download_ready"
            fallback["candidate_status"] = "PARTIAL_SUCCESS" if fallback.get("hold_records") else "SUCCESS"
            log("[awb] Download button stayed enabled; proceeding without waiting for optional result wording")
            return fallback
        if result["state"] == "failed_no_download":
            save_page_artifacts(page, screenshot_dir, "awb_failed_no_download")
            raise AWBModalTimeoutError("AWB_MODAL_TIMEOUT: AWB modal failed without an enabled Download button")
        page.wait_for_timeout(1000)
    save_page_artifacts(page, screenshot_dir, "awb_modal_timeout")
    text = re.sub(r"\s+", " ", last_snapshot.get("text", ""))[:300]
    raise AWBModalTimeoutError(f"AWB_MODAL_TIMEOUT: Download button did not become enabled within 60 seconds. Last modal text: {text}")

def click_last_download(page, action_timeout: int, deadline: float, account: str, start: float, hb: Dict[str, float]) -> None:
    end_time = min(time.monotonic() + action_timeout, deadline)
    button = page.locator("button:has-text('Download')").last
    button.wait_for(state="visible", timeout=max(1000, int((end_time - time.monotonic()) * 1000)))
    while time.monotonic() < end_time:
        heartbeat(account, "下载ZIP", start, hb)
        try:
            if button.is_visible() and button.is_enabled():
                button.click(timeout=min(action_timeout * 1000, 5000))
                return
        except Exception:
            pass
        page.wait_for_timeout(300)
    raise TimeoutError("Download button did not become enabled")



def looks_like_zip_response(response) -> bool:
    try:
        headers = {k.lower(): v.lower() for k, v in response.headers.items()}
        ctype = headers.get("content-type", "")
        disposition = headers.get("content-disposition", "")
        url = response.url.lower()
        return "zip" in ctype or "octet-stream" in ctype or "zip" in disposition or "awb" in url or ("air" in url and "waybill" in url)
    except Exception:
        return False


def save_response_zip(response, target: Path) -> Path:
    body = response.body()
    if not body.startswith(b"PK"):
        raise RuntimeError(f"Download response was not a ZIP file: {response.url}")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(body)
    return target


def download_awb_zip(page, dirs: Dict[str, Path], account: str, action_timeout: int) -> Path:
    captured_responses = []
    captured_downloads = []
    diagnostics = {"requests": [], "responses": [], "downloads": [], "popups": [], "new_pages": [], "trace": "", "zip_response_saved": "", "json_download_url": "", "default_zip_detected": "", "click_count": 0}
    start_time = time.time()
    before_zips = set(dirs["downloads"].glob("*.zip"))

    def on_request(request):
        try:
            diagnostics["requests"].append({"method": request.method, "url": request.url, "resource_type": request.resource_type})
        except Exception:
            pass

    def on_response(response):
        try:
            headers = {k.lower(): v for k, v in response.headers.items()}
            diagnostics["responses"].append({"status": response.status, "url": response.url, "content_type": headers.get("content-type", ""), "content_disposition": headers.get("content-disposition", "")})
            if looks_like_zip_response(response):
                captured_responses.append(response)
            url = extract_download_url_from_json_response(response)
            if url:
                diagnostics["json_download_url"] = url
        except Exception:
            pass

    def on_download(download):
        captured_downloads.append(download)
        try:
            diagnostics["downloads"].append({"suggested_filename": download.suggested_filename})
        except Exception:
            pass

    def on_popup(popup):
        try:
            diagnostics["popups"].append({"url": popup.url})
        except Exception:
            diagnostics["popups"].append({"url": ""})

    def on_page(new_page):
        try:
            diagnostics["new_pages"].append({"url": new_page.url})
        except Exception:
            diagnostics["new_pages"].append({"url": ""})

    context = getattr(page, "context", None)
    trace_path = unique_path(dirs["screenshots"] / f"{account}_playwright_trace.zip")
    tracing_started = False
    try:
        page.on("request", on_request)
        page.on("response", on_response)
        page.on("download", on_download)
        page.on("popup", on_popup)
        if context is not None:
            try:
                context.on("page", on_page)
                context.tracing.start(screenshots=True, snapshots=True, sources=False)
                tracing_started = True
            except Exception:
                pass

        # One and only one modal Download click. The real site normally returns JSON
        # containing a signed Google Storage URL, so standard expect_download is not
        # used as the primary path.
        click_awb_popup_download(page, action_timeout)
        diagnostics["click_count"] = 1
        end_time = time.monotonic() + max(action_timeout, 30)
        attempted_urls = set()
        attempted_responses = set()
        while time.monotonic() < end_time:
            if captured_downloads:
                download = captured_downloads.pop(0)
                suggested = download.suggested_filename or f"{account}_AWB.zip"
                if not suggested.lower().endswith(".zip"):
                    suggested += ".zip"
                zip_path = unique_path(dirs["downloads"] / suggested)
                download.save_as(zip_path)
                if zipfile.is_zipfile(zip_path):
                    return zip_path

            url = diagnostics.get("json_download_url") or ""
            if url and url not in attempted_urls:
                attempted_urls.add(url)
                zip_path = unique_path(dirs["downloads"] / f"{account}_AWB_json_url.zip")
                fetcher = None
                if context is not None and getattr(context, "request", None) is not None:
                    def fetcher(value):
                        response = context.request.get(value, timeout=max(action_timeout, 30) * 1000)
                        if not response.ok:
                            raise RuntimeError(f"Signed URL request failed: {response.status}")
                        return response.body()
                saved = save_zip_from_url(url, zip_path, fetcher=fetcher)
                diagnostics["zip_response_saved"] = str(saved)
                return saved

            for response in list(captured_responses):
                key = getattr(response, "url", str(id(response)))
                if key in attempted_responses:
                    continue
                attempted_responses.add(key)
                try:
                    zip_path = unique_path(dirs["downloads"] / f"{account}_AWB_response.zip")
                    saved = save_response_zip(response, zip_path)
                    diagnostics["zip_response_saved"] = str(saved)
                    return saved
                except Exception:
                    pass

            detected = find_new_zip_after_click(dirs["downloads"], before_zips, start_time)
            if detected and zipfile.is_zipfile(detected):
                diagnostics["default_zip_detected"] = str(detected)
                return detected
            page.wait_for_timeout(250)

        raise DownloadCaptureFailed("DOWNLOAD_CAPTURE_FAILED: no ZIP was captured after the single Download click")
    finally:
        if tracing_started and context is not None:
            try:
                context.tracing.stop(path=str(trace_path))
                diagnostics["trace"] = str(trace_path)
            except Exception:
                pass
        try:
            write_download_diagnostics(diagnostics, dirs, account)
        except Exception:
            pass
        for event, handler in [("request", on_request), ("response", on_response), ("download", on_download), ("popup", on_popup)]:
            try:
                page.remove_listener(event, handler)
            except Exception:
                pass
        if context is not None:
            try:
                context.remove_listener("page", on_page)
            except Exception:
                pass

def unzip_labels(zip_path: Path, original_dir: Path) -> List[Path]:
    extracted: List[Path] = []
    with zipfile.ZipFile(zip_path) as archive:
        for item in archive.namelist():
            if not item.lower().endswith(".pdf"):
                continue
            target = unique_path(original_dir / make_new_name(Path(item).name))
            with archive.open(item) as source:
                target.write_bytes(source.read())
            extracted.append(target)
    return extracted


def process_zip_records(zip_path: Path, account: str, run_id: str, dirs: Dict[str, Path] | None = None) -> Tuple[List[Dict], Dict]:
    zip_path = Path(zip_path).resolve()
    if not zip_path.exists() or zip_path.stat().st_size <= 0:
        raise FileNotFoundError(f"ZIP missing or empty: {zip_path}")
    if not zipfile.is_zipfile(zip_path):
        raise RuntimeError(f"ZIP_CORRUPTED: invalid ZIP file: {zip_path}")
    with zipfile.ZipFile(zip_path) as archive:
        bad = archive.testzip()
        if bad:
            raise RuntimeError(f"ZIP_CORRUPTED: first bad entry: {bad}")
        pdf_entries = [name for name in archive.namelist() if name.lower().endswith(".pdf")]
    dirs = dirs or account_run_dirs(account, run_id)
    extracted = unzip_labels(zip_path, dirs["original"])
    log(f"[pdf] 解压PDF数量 {len(extracted)}")
    records = process_pdf_list(extracted, dirs["processed"], dirs["unknown"], dirs["failed"], ADDRESS_FILE, account=account, run_id=run_id)
    for record in records:
        record.update({"account": account, "date": datetime.now().strftime("%Y-%m-%d"), "run_id": run_id})
    summary = {
        "zip_path": str(zip_path),
        "zip_size": zip_path.stat().st_size,
        "zip_integrity": "OK",
        "zip_pdf_entries": len(pdf_entries),
        "extracted_pdf_count": len(extracted),
        "processed_success_count": sum(1 for r in records if r.get("status") == "SUCCESS"),
        "processed_count": sum(1 for r in records if r.get("status") in {"SUCCESS", "UNKNOWN"}),
        "duplicate_skipped_count": sum(1 for r in records if r.get("status") == "DUPLICATE_SKIPPED"),
    }
    return records, summary


def select_to_ship(page, config: Dict, action_timeout: int) -> None:
    selectors = config.get("selectors", {})
    status_placeholder = selectors.get("status_placeholder", "#react-select-4-placeholder")
    to_ship_option = selectors.get("to_ship_option", "#react-select-4-option-2")
    timeout_ms = action_timeout * 1000
    try:
        page.locator(status_placeholder).click(timeout=timeout_ms)
        page.wait_for_timeout(500)
        page.locator(to_ship_option).click(timeout=timeout_ms)
    except Exception:
        page.get_by_text("Status", exact=True).click(timeout=timeout_ms)
        page.wait_for_timeout(300)
        page.get_by_text("To Ship", exact=True).click(timeout=timeout_ms)
    page.wait_for_timeout(1500)


def run_live_account(playwright, account: Dict, config: Dict, run_id: str, account_timeout: int, action_timeout: int) -> Tuple[List[Dict], List[Dict]]:
    name = account["name"]
    dirs = account_run_dirs(name, run_id)
    state_file = BASE_DIR / account["state"]
    records: List[Dict] = []
    errors: List[Dict] = []
    browser = None
    context = None
    page = None
    start = time.monotonic()
    deadline = start + account_timeout
    hb = {"last": 0.0}

    try:
        stage(name, "启动账号", start, hb)
        browser = playwright.chromium.launch(headless=bool(config.get("headless", False)), downloads_path=str(dirs["downloads"]))
        context = browser.new_context(storage_state=str(state_file), accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(action_timeout * 1000)

        stage(name, "打开页面", start, hb)
        check_deadline(deadline, name, "打开页面")
        page.goto(config["sales_url"], wait_until="domcontentloaded", timeout=action_timeout * 1000)
        page.wait_for_timeout(1500)

        stage(name, "检查登录", start, hb)
        text = get_page_text(page).lower()
        if "login" in text and "password" in text:
            raise RuntimeError("LOGIN_EXPIRED: login state expired; run save_login.py for this account")

        stage(name, "检查订单", start, hb)
        select_to_ship(page, config, action_timeout)
        if looks_like_no_orders(get_page_text(page)) or not has_selectable_orders(page):
            return [no_orders_record(name, run_id)], errors

        awb_button = find_visible_text_locator(page, "AWB Download")
        if awb_button is None or locator_has_disabled_state(awb_button):
            return [no_orders_record(name, run_id)], errors
        awb_button.click(timeout=action_timeout * 1000)
        page.wait_for_timeout(800)

        download_state = evaluate_to_ship_download_state(page)
        if download_state == "no_orders":
            return [no_orders_record(name, run_id)], errors
        download_all = find_visible_text_locator(page, "Download All")
        if download_all is None or locator_has_disabled_state(download_all):
            return [no_orders_record(name, run_id)], errors

        stage(name, "生成AWB", start, hb)
        download_all.click(timeout=action_timeout * 1000)
        page.wait_for_selector("text=Bulk downloading Air Waybill", timeout=action_timeout * 1000)
        log("[awb] AWB弹窗出现")
        generation_result = wait_for_awb_result(page, dirs["screenshots"], action_timeout, deadline, name, start, hb, run_id)
        hold_records = generation_result.get("hold_records", [])
        if hold_records:
            errors.extend(hold_records)
        log(f"{name} AWB generation: success_count={generation_result.get('success_count', 0)} hold_count={len(hold_records)}")

        stage(name, "下载ZIP", start, hb)
        try:
            zip_path = download_awb_zip(page, dirs, name, action_timeout)
        except DownloadCaptureFailed as exc:
            save_page_artifacts(page, dirs["screenshots"], "zip_download_capture_failed")
            errors.append(error_record(name, run_id, "DOWNLOAD_CAPTURE_FAILED", str(exc)))
            stage(name, "账号完成", start, hb)
            return records, errors
        except Exception as exc:
            save_page_artifacts(page, dirs["screenshots"], "zip_download_failed")
            errors.append(error_record(name, run_id, "DOWNLOAD_FAILED", str(exc)))
            stage(name, "账号完成", start, hb)
            return records, errors
        log(f"[download] ZIP保存完成 {zip_path}")
        log(f"[download] ZIP文件大小 {zip_path.stat().st_size}")

        stage(name, "处理本地ZIP", start, hb)
        records, zip_summary = process_zip_records(zip_path, name, run_id, dirs)
        log(f"[pdf] 处理PDF数量 {zip_summary['processed_count']}")
        stage(name, "账号完成", start, hb)
        return records, errors

    except AWBModalTimeoutError as exc:
        if page is not None:
            save_page_artifacts(page, dirs["screenshots"], "awb_modal_timeout")
        errors.append(error_record(name, run_id, "AWB_MODAL_TIMEOUT", str(exc)))
        return records, errors
    except AccountTimeoutError as exc:
        if page is not None:
            save_page_artifacts(page, dirs["screenshots"], "account_timeout")
        errors.append(error_record(name, run_id, "ACCOUNT_TIMEOUT", str(exc)))
        return records, errors
    except Exception as exc:
        if page is not None:
            try:
                save_page_artifacts(page, dirs["screenshots"], "account_error")
            except Exception:
                pass
        message = str(exc)
        status = "LOGIN_EXPIRED" if "LOGIN_EXPIRED" in message else "ACCOUNT_FAILED"
        errors.append(error_record(name, run_id, status, message))
        return records, errors
    finally:
        for obj in [page, context, browser]:
            if obj is not None:
                try:
                    obj.close()
                except Exception:
                    pass


def create_demo_pdfs(folder: Path) -> None:
    try:
        import pymupdf as fitz
    except ImportError:
        import fitz
    folder.mkdir(parents=True, exist_ok=True)
    samples = [
        ("DEMO-SHOE US 8.5 90990001.pdf", [(25, 28, "OLD SENDER ADDRESS"), (25, 55, "OLD CITY STATE 00000"), (25, 150, "UPS GROUND"), (25, 195, "TRACKING #: 1Z999AA10123456784")]),
        ("DEMO-TEE XL 90990002.pdf", [(25, 90, "USPS GROUND ADVANTAGE"), (25, 135, "9400111899223856928499"), (28, 260, "KICKS CREW SNEAKERS")]),
        ("DEMO-HAT M 90990003.pdf", [(25, 45, "FedEx"), (25, 90, "TRK# 8732 0160 7478"), (25, 135, "INTL ECONOMY"), (25, 180, "ORIGIN ID"), (25, 225, "BILL SENDER")]),
    ]
    for filename, lines in samples:
        doc = fitz.open()
        page = doc.new_page(width=612, height=792)
        for x, y, line in lines:
            page.insert_text((x, y), line, fontsize=16)
        page.insert_text((25, 360), "RECIPIENT AREA - DO NOT MODIFY", fontsize=14)
        doc.save(folder / filename)
        doc.close()


def run_demo() -> Tuple[List[Dict], Path, Path]:
    account = "DEMO_ACCOUNT"
    run_id = generate_run_id()
    dirs = account_run_dirs(account, run_id)
    create_demo_pdfs(dirs["original"])
    records = process_folder(dirs["original"], dirs["processed"], dirs["unknown"], dirs["failed"], ADDRESS_FILE, account=account, run_id=run_id)
    for record in records:
        record.update({"account": account, "date": datetime.now().strftime("%Y-%m-%d"), "run_id": run_id})
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report = create_report(records, REPORT_DIR / f"今日发货_DEMO_{stamp}.xlsx")
    archive = zip_processed(records, REPORT_DIR / f"处理后PDF_DEMO_{stamp}.zip")
    return records, report, archive


COMPLETED_ACCOUNT_STATES = {"SUCCESS", "PARTIAL_SUCCESS", "NO_ORDERS"}


def latest_today_report() -> Path | None:
    today = datetime.now().strftime("%Y%m%d")
    reports = [p for p in REPORT_DIR.glob(f"*{today}*.xlsx") if p.is_file() and not p.name.startswith("~$")]
    if not reports:
        return None
    return max(reports, key=lambda p: p.stat().st_mtime)


def read_account_summary(report: Path | None = None) -> Dict[str, str]:
    report = report or latest_today_report()
    if report is None or not report.exists():
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(report, read_only=True, data_only=True)
        if len(wb.worksheets) < 3:
            wb.close()
            return {}
        result = {}
        for row in wb.worksheets[2].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                result[str(row[0])] = str(row[1] or "")
        wb.close()
        return result
    except Exception:
        return {}


def incomplete_account_names(config: Dict, report: Path | None = None) -> Tuple[List[str], Dict[str, str]]:
    summary = read_account_summary(report)
    names = [account["name"] for account in config.get("accounts", [])]
    incomplete = [name for name in names if summary.get(name) not in COMPLETED_ACCOUNT_STATES]
    return incomplete, summary


def records_from_report(report: Path | None) -> List[Dict]:
    if report is None or not report.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(report, read_only=True, data_only=True)
        records: List[Dict] = []
        if len(wb.worksheets) >= 1:
            for row in wb.worksheets[0].iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                records.append({
                    "date": str(row[0] or ""),
                    "run_id": str(row[1] or ""),
                    "account": str(row[2] or ""),
                    "sale_id": str(row[3] or ""),
                    "sku": str(row[4] or ""),
                    "size": str(row[5] or ""),
                    "quantity": str(row[6] or ""),
                    "carrier": str(row[7] or ""),
                    "tracking_number": str(row[8] or ""),
                    "source_file": str(row[9] or ""),
                    "output_file": str(row[10] or ""),
                    "status": str(row[11] or ""),
                    "message": str(row[12] or ""),
                    "filename": Path(str(row[10] or row[9] or "")).name,
                })
        if len(wb.worksheets) >= 2:
            for row in wb.worksheets[1].iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                records.append({
                    "date": str(row[0] or ""),
                    "run_id": str(row[1] or ""),
                    "account": str(row[2] or ""),
                    "sale_id": str(row[3] or ""),
                    "filename": str(row[4] or ""),
                    "status": str(row[5] or ""),
                    "error": str(row[6] or ""),
                    "source_file": str(row[7] or ""),
                })
        wb.close()
        return records
    except Exception as exc:
        log(f"[rerun] previous report could not be merged: {exc}")
        return []


def generate_recovery_run_id() -> str:
    return f"recovery_{datetime.now():%Y%m%d_%H%M%S}"


def run_resume_zip(zip_path: Path, account: str, hold_sale_id: str = "", hold_error: str = "Order on hold", run_id: str | None = None) -> Tuple[List[Dict], Path, Path, Dict]:
    if not account:
        raise ValueError("--account is required with --resume-zip")
    zip_path = zip_path.resolve()
    run_id = run_id or generate_recovery_run_id()
    dirs = account_run_dirs(account, run_id)
    recovered_zip = unique_path(dirs["downloads"] / zip_path.name)
    shutil.copy2(zip_path, recovered_zip)
    log(f"[recovery] ZIP路径 {zip_path}")
    records, summary = process_zip_records(recovered_zip, account, run_id, dirs)
    errors: List[Dict] = []
    if hold_sale_id:
        errors.append(order_on_hold_record(account, run_id, hold_sale_id, hold_error or "Order on hold"))
    all_records = records + errors
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report = create_report(all_records, REPORT_DIR / f"今日发货_RECOVERY_{stamp}.xlsx")
    archive = zip_processed(all_records, REPORT_DIR / f"处理后PDF_RECOVERY_{stamp}.zip")
    summary.update({
        "run_id": run_id,
        "hold_count": len(errors),
        "report": str(report),
        "archive": str(archive),
    })
    status_extra = {k: v for k, v in summary.items() if k not in {"run_id", "report", "archive"}}
    carriers = Counter((r.get("carrier") or "").upper() for r in records if r.get("status") in {"SUCCESS", "UNKNOWN"})
    write_status(
        "completed", "Offline ZIP recovery completed", run_id=run_id, report=str(report), archive=str(archive),
        total=len(all_records), success_pdf_count=summary["processed_count"], successful_pdfs=summary["processed_count"],
        exception_count=len(errors), exceptions=len(errors), ups_count=carriers["UPS"], usps_count=carriers["USPS"],
        fedex_count=carriers["FEDEX"], unknown_count=carriers["UNKNOWN"], account_statuses={account: "PARTIAL_SUCCESS" if records and errors else "SUCCESS" if records else "FAILED"},
        current_account=account, current_step="已完成", **status_extra
    )
    log(f"[recovery] Excel路径 {report}")
    log(f"[recovery] PDF ZIP路径 {archive}")
    return all_records, report, archive, summary

def main() -> int:
    parser = argparse.ArgumentParser(description="CrewSupply download + PDF processing + daily report")
    parser.add_argument("--all", action="store_true", help="Run all configured accounts")
    parser.add_argument("--account", help="Run one configured account")
    parser.add_argument("--process-only", help="Process PDFs in a local folder")
    parser.add_argument("--resume-zip", help="Offline recovery from an existing AWB ZIP; never starts browser or network")
    parser.add_argument("--hold-sale-id", help="Sale ID to record as ORDER_ON_HOLD during --resume-zip")
    parser.add_argument("--hold-error", default="Order on hold", help="Hold error text for --resume-zip")
    parser.add_argument("--demo", action="store_true", help="Create demo labels and run the full local pipeline")
    parser.add_argument("--account-timeout", type=int, default=120, help="Max seconds per account")
    parser.add_argument("--action-timeout", type=int, default=20, help="Max seconds per browser action")
    parser.add_argument("--production-run", action="store_true", help="Formal daily production run; does not consume test real-run quota")
    parser.add_argument("--rerun-incomplete", action="store_true", help="Run only accounts not completed in today latest report and merge results")
    args = parser.parse_args()

    run_id = generate_recovery_run_id() if args.resume_zip else generate_run_id()
    write_status("running", "Task started", run_id=run_id, current_account="", current_step="启动中", successful_pdfs=0, success_pdf_count=0, exceptions=0, exception_count=0, duplicate_skipped_count=0, ups_count=0, usps_count=0, fedex_count=0, unknown_count=0, account_statuses={}, report="", archive="")
    try:
        if args.resume_zip:
            records, report, archive, summary = run_resume_zip(Path(args.resume_zip), args.account or "", args.hold_sale_id or "", args.hold_error, run_id=run_id)
            return 0

        if args.demo:
            records, report, archive = run_demo()
            write_status("completed", "Demo completed", report=str(report), archive=str(archive), total=len(records), run_id=run_id)
            log(f"Report: {report}")
            log(f"PDF ZIP: {archive}")
            return 0

        config = load_config()
        if args.process_only:
            source = Path(args.process_only).resolve()
            account = source.name or "LOCAL"
            dirs = account_run_dirs(account, run_id)
            records = process_folder(source, dirs["processed"], dirs["unknown"], dirs["failed"], ADDRESS_FILE, account=account, run_id=run_id, run_date=datetime.now().strftime("%Y-%m-%d"))
            for record in records:
                record.update({"account": account, "date": datetime.now().strftime("%Y-%m-%d"), "run_id": run_id})
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report = create_report(records, REPORT_DIR / f"今日发货_{stamp}.xlsx")
            archive = zip_processed(records, REPORT_DIR / f"处理后PDF_{stamp}.zip")
            write_status("completed", "Local PDF processing completed", report=str(report), archive=str(archive), total=len(records), run_id=run_id)
            return 0

        selected = config["accounts"]
        previous_report = latest_today_report() if args.rerun_incomplete else None
        previous_records: List[Dict] = records_from_report(previous_report) if args.rerun_incomplete else []
        if args.rerun_incomplete:
            incomplete, summary = incomplete_account_names(config, previous_report)
            log(f"[rerun] latest_report={previous_report}")
            log(f"[rerun] account_summary={summary}")
            log(f"[rerun] selected_accounts={incomplete}")
            selected = [a for a in selected if a["name"] in set(incomplete)]
            if not selected:
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report = create_report(previous_records, REPORT_DIR / f"今日发货_RERUN_{stamp}.xlsx")
                archive = zip_processed(previous_records, REPORT_DIR / f"处理后PDF_RERUN_{stamp}.zip")
                write_status("completed", "No incomplete accounts to rerun", report=str(report), archive=str(archive), run_id=run_id, total=len(previous_records), account_statuses=summary)
                return 0
        elif args.account:
            selected = [a for a in selected if a["name"] == args.account]
            if not selected:
                raise ValueError(f"Unknown account: {args.account}")
        elif not args.all:
            parser.error("Choose --all, --account, --process-only, --rerun-incomplete or --demo")

        scope = "rerun_incomplete" if args.rerun_incomplete else args.account if args.account else "all"
        if not args.production_run:
            if not can_start_real_run():
                raise RuntimeError("REAL_RUN_LIMIT_REACHED: maximum two authorized real runs already used")
            record_real_run(scope, " ".join(sys.argv))
        else:
            log("[production] formal production run enabled; test real-run quota is not modified")
        from playwright.sync_api import sync_playwright
        all_records: List[Dict] = list(previous_records) if args.rerun_incomplete else []
        all_errors: List[Dict] = []
        with sync_playwright() as playwright:
            for index, account in enumerate(selected, start=1):
                write_status("running", f"Processing {account['name']}", current=index, total_accounts=len(selected), run_id=run_id)
                records, errors = run_live_account(playwright, account, config, run_id, args.account_timeout, args.action_timeout)
                if not records and not errors:
                    errors = [error_record(account["name"], run_id, "ACCOUNT_FAILED", "Account produced no result")]
                all_records.extend(records)
                all_errors.extend(errors)

        all_records.extend(all_errors)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stage_start = time.monotonic()
        update_status(current_account="", current_step="生成报表")
        log("[REPORT] 生成报表")
        report = create_report(all_records, REPORT_DIR / f"今日发货_{stamp}.xlsx")
        log(f"[report] Excel生成完成 {report}")
        archive = zip_processed(all_records, REPORT_DIR / f"处理后PDF_{stamp}.zip")
        shipping = [r for r in all_records if r.get("status") in {"SUCCESS", "UNKNOWN"}]
        exceptions = [r for r in all_records if r.get("status") in {"FAILED", "CHECK", "PDF_PROCESS_FAILED", "ACCOUNT_TIMEOUT", "ACCOUNT_FAILED", "LOGIN_EXPIRED", "DOWNLOAD_FAILED", "DOWNLOAD_CAPTURE_FAILED", "AWB_GENERATION_FAILED", "AWB_MODAL_TIMEOUT", "ORDER_ON_HOLD"}]
        carriers = Counter((r.get("carrier") or "").upper() for r in shipping)
        by_account = defaultdict(list)
        for record in all_records:
            by_account[record.get("account", "")].append(record)
        from shipping_report import account_state
        account_statuses = {name: account_state(items) for name, items in by_account.items() if name}
        no_order_count = sum(status == "NO_ORDERS" for status in account_statuses.values())
        write_status(
            "completed", "All accounts completed", report=str(report), archive=str(archive), run_id=run_id,
            total=len(all_records), success_pdf_count=len(shipping), successful_pdfs=len(shipping),
            exception_count=len(exceptions), exceptions=len(exceptions), no_order_count=no_order_count,
            duplicate_skipped_count=sum(r.get("status") == "DUPLICATE_SKIPPED" for r in all_records),
            ups_count=carriers["UPS"], usps_count=carriers["USPS"], fedex_count=carriers["FEDEX"], unknown_count=carriers["UNKNOWN"],
            account_statuses=account_statuses, completed_accounts=len(account_statuses), current_account="", current_step="已完成",
            elapsed_seconds=int(time.monotonic() - stage_start)
        )
        return 0
    except Exception as exc:
        message = str(exc)
        precheck_markers = ["Unexpected UTF-8 BOM", "REAL_RUN_LIMIT_REACHED", "Unknown account", "Choose --all"]
        state = "PRECHECK_FAILED" if any(marker in message for marker in precheck_markers) else "failed"
        write_status(state, message, run_id=run_id)
        print(exc, file=sys.stderr, flush=True)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

























